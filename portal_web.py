# =============================================================================
# portal_web.py
# Motor Web y Enrutador Principal — Portal B2B Importadora Uziel
# Proyecto: Dashboard de Marketing - Importadora Uziel C.A.
# Autor: Jesús | Framework: Flask 3.0.0 | Python 3.11.8
# =============================================================================
#
# DESCRIPCIÓN:
#   Este archivo es el servidor Flask. Recibe peticiones HTTP del navegador,
#   verifica permisos, consulta la base de datos y devuelve el HTML correcto.
#
#   Rutas disponibles:
#     GET/POST /login              — Pantalla de inicio de sesión
#     GET      /logout             — Cierra la sesión del usuario
#     GET      /                   — Dashboard principal con estadísticas y tareas
#     GET      /catalogo           — Módulo PIM: lista completa de productos con miniaturas
#     GET      /producto/<sku>     — Ficha de detalle del producto con galería DAM
#     GET      /activos/<archivo>  — Sirve archivos estáticos de almacen_activos/
#     GET      /clientes           — Módulo CRM: directorio de clientes
#     GET/POST /nuevo_producto     — Formulario para agregar producto (solo Admin)
#     GET/POST /nuevo_cliente      — Formulario para agregar cliente (solo Admin)
#     GET/POST /editar/<sku>       — Formulario para editar producto (solo Admin)
#     GET      /eliminar/<sku>     — Elimina un producto (solo Admin)
#     POST     /generar_pdf        — Genera y descarga un PDF con productos seleccionados
#     GET      /tareas             — Gestión de tareas (todas para Admin, propias para otros)
#     POST     /asignar_tarea      — Crea una nueva tarea (solo Admin)
#     POST     /actualizar_tarea/<id>   — Cambia el estado de una tarea
#     GET      /admin/usuarios           — Panel de gestión de usuarios (solo Admin)
#     POST     /admin/usuario/nuevo      — Crear un nuevo usuario (solo Admin)
#     POST     /admin/usuario/<u>/editar — Actualizar datos y permisos (solo Admin)
#     POST     /admin/usuario/<u>/pass   — Cambiar contraseña de usuario (solo Admin)
#     POST     /admin/usuario/<u>/borrar — Eliminar usuario (solo Admin)
#     GET      /cliente/<rif>           — Ficha completa del cliente con historial
#     POST     /cliente/<rif>/editar    — Actualizar datos de contacto del cliente
#     GET      /cotizaciones            — Lista todas las cotizaciones
#     GET/POST /cotizacion/nueva        — Crear nueva cotización (acepta ?cliente_rif=)
#     GET      /cotizacion/<id>         — Detalle de una cotización
#     POST     /cotizacion/<id>/estado  — Cambiar estado de cotización
#     GET      /cotizacion/<id>/pdf     — Descargar PDF de cotización
#
# DESPLIEGUE EN RENDER:
#   Build Command : pip install -r requirements.txt
#   Start Command : gunicorn portal_web:app
# =============================================================================

import os
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, send_from_directory, abort, make_response
from functools import wraps
import io
import secrets
from PIL import Image
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas as pdf_canvas
from src.database import ConexionBD
from src.generador_pdf import generar_pdf_alianza, generar_ficha_tecnica, generar_pdf_catalogo, generar_reporte_pdf
from src.backup import iniciar_hilo_respaldos, crear_respaldo

# Iniciar hilo de respaldos automáticos
try:
    iniciar_hilo_respaldos()
except Exception as e:
    print("Error iniciando hilo de respaldos:", e)

# =============================================================================
# ██████████████ CONFIGURACIÓN - EDITAR AQUÍ ██████████████
# =============================================================================

# Clave secreta para cifrar las cookies de sesión.
#
# CÓMO CONFIGURAR (elige una opción):
#
#   Opción A — Variable de entorno (RECOMENDADO para producción):
#     Windows CMD : set FLASK_SECRET_KEY=tu_clave_larga_aqui
#     Windows PS  : $env:FLASK_SECRET_KEY="tu_clave_larga_aqui"
#     Linux/Mac   : export FLASK_SECRET_KEY="tu_clave_larga_aqui"
#     Render.com  : Dashboard > Environment > Add Environment Variable
#
#   Genera una clave segura con:
#     python -c "import secrets; print(secrets.token_hex(32))"
#
#   Opción B — Reemplaza la cadena de abajo (solo para desarrollo local,
#              NUNCA subas esto a GitHub):
#
SECRET_KEY_DEFAULT = "cambia_esta_clave_antes_de_produccion"
SECRET_KEY = os.getenv("FLASK_SECRET_KEY", SECRET_KEY_DEFAULT)

# Advertencia si se usa la clave por defecto fuera de desarrollo local
if SECRET_KEY == SECRET_KEY_DEFAULT:
    import warnings
    warnings.warn(
        "\n  SEGURIDAD: Se está usando la SECRET_KEY por defecto.\n"
        "   Configura la variable de entorno FLASK_SECRET_KEY antes de ir a producción.\n"
        "   Genera una clave segura con: python -c \"import secrets; print(secrets.token_hex(32))\"",
        stacklevel=1
    )

# Clave API para que la aplicación de escritorio pueda subir imágenes.
# Debe ser la misma en main.py (variable API_KEY).
# En Render, configúrala como variable de entorno: DESKTOP_API_KEY
API_KEY_DESKTOP = os.getenv("DESKTOP_API_KEY", "uziel-desktop-sync-2026")

# Nombre de la empresa que aparece en el encabezado del PDF generado
NOMBRE_EMPRESA_PDF = "Catálogo de Productos - Importadora Uziel"
SUBTITULO_PDF = "Generado automáticamente desde el Portal B2B"

# Nombre del archivo PDF descargado por el usuario
NOMBRE_ARCHIVO_PDF = "Catalogo_Uziel.pdf"

import re

def validar_complejidad_password(password):
    """
    Valida que la contraseña cumpla con los requisitos mínimos de seguridad:
    - Al menos 8 caracteres de longitud.
    - Al menos una letra mayúscula.
    - Al menos una letra minúscula.
    - Al menos un número.
    - Al menos un carácter especial.
    """
    if len(password) < 8:
        return False
    if not re.search(r'[A-Z]', password):
        return False
    if not re.search(r'[a-z]', password):
        return False
    if not re.search(r'[0-9]', password):
        return False
    if not re.search(r'[^A-Za-z0-9]', password):
        return False
    return True

# =============================================================================

app = Flask(__name__)
app.secret_key = SECRET_KEY

# Configurar propiedades seguras para cookies de sesión
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=True
)

# Instancia global del módulo de base de datos
bd = ConexionBD()

# =============================================================================
# PROTECCIÓN CSRF MANUAL
# =============================================================================

@app.before_request
def asegurar_csrf_token():
    """Genera un token CSRF si no existe en la sesión."""
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)

@app.before_request
def validar_csrf_token():
    """Valida el token CSRF en peticiones POST, excepto para endpoints de la API del Desktop."""
    if request.method == 'POST':
        # Los endpoints del Desktop (/api/...) usan X-API-Key en vez de cookies de sesión
        if request.path.startswith('/api/'):
            return
        
        token = request.form.get('csrf_token') or request.headers.get('X-CSRF-Token')
        expected = session.get('csrf_token')
        
        if not expected or token != expected:
            print(f"[CSRF] Validación fallida para ruta: {request.path}")
            abort(400, "Token CSRF inválido o ausente.")

@app.before_request
def validar_password_confirmacion():
    """Reclama la contraseña para cualquier acción de modificación en peticiones POST (web)."""
    if request.method == 'POST':
        # Ignorar rutas que no requieren contraseña (login, generar_pdf, APIs)
        if request.path.startswith('/api/'):
            return
        if request.path == '/login' or request.path == '/generar_pdf':
            return
        
        username = session.get('usuario')
        if not username:
            return  # Dejar que el decorador login_requerido o la lógica de la ruta lo maneje
            
        pwd = request.form.get('verificar_password')
        if not pwd:
            flash(' Por seguridad, debes ingresar tu contraseña para confirmar esta acción.', 'error')
            return redirect(request.referrer or url_for('inicio'))
            
        if not bd.verificar_contrasena_usuario(username, pwd):
            flash(' La contraseña de verificación es incorrecta. Acción denegada.', 'error')
            return redirect(request.referrer or url_for('inicio'))

@app.context_processor
def inyectar_csrf():
    """Inyecta el token CSRF en el contexto de todos los templates."""
    return {'csrf_token': session.get('csrf_token', '')}


@app.context_processor
def inyectar_datos_vzla():
    """Inyecta la base de datos de estados y municipios de Venezuela en las plantillas."""
    import json
    try:
        ruta_vzla = os.path.join(os.path.dirname(__file__), 'src', 'venezuela.json')
        with open(ruta_vzla, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return {'venezuela_data_js': json.dumps(data)}
    except Exception as e:
        print(f"Error al cargar venezuela.json: {e}")
        return {'venezuela_data_js': '{}'}



# =============================================================================
# UTILIDAD — Extraer ruta relativa a almacen_activos/
# =============================================================================

def _extraer_ruta_relativa(ruta_archivo):
    """
    Convierte cualquier formato de ruta (absoluta Windows/Linux o relativa)
    a una ruta relativa dentro de almacen_activos/.

    Ejemplos:
      'G:/.../almacen_activos/16572-0P030/1.jpg'  → '16572-0P030/1.jpg'
      '/opt/render/.../almacen_activos/SKU/2.jpg'  → 'SKU/2.jpg'
      '16572-0P030/1.jpg'                           → '16572-0P030/1.jpg'
      '1.jpg'                                       → '1.jpg'  (caso borde)
    """
    if not ruta_archivo:
        return None
    ruta = ruta_archivo.replace('\\', '/')
    # Si contiene 'almacen_activos/', extraer lo que sigue
    if 'almacen_activos/' in ruta:
        idx = ruta.index('almacen_activos/')
        return ruta[idx + len('almacen_activos/'):]
    # Si es ruta absoluta (Windows con letra de unidad o Linux con /)
    if ruta.startswith('/') or (len(ruta) > 2 and ruta[1] == ':'):
        return os.path.basename(ruta)
    # Ya es relativa
    return ruta

# Crear tablas automáticamente al arrancar el servidor —
# operaciones idempotentes, seguras de ejecutar en cada inicio
bd.inicializar_tareas()
bd.inicializar_cotizaciones()
bd.inicializar_permisos_usuarios()
bd.inicializar_gastos()


# =============================================================================
# CONTEXTO GLOBAL — disponible en TODOS los templates automáticamente
# =============================================================================

@app.context_processor
def inyectar_notificaciones():
    """
    Inyecta el conteo de tareas pendientes del usuario logueado en todos
    los templates. Así cualquier página puede mostrar el badge de campana
    sin que cada ruta tenga que calcularlo por separado.

    La variable 'total_notif' queda disponible en todos los Jinja2 templates.
    Además inyecta 'puede_ver(modulo)' para controlar la visibilidad del sidebar.
    """
    # Si no hay sesión activa, no hay notificaciones que mostrar
    if 'usuario' not in session:
        return {'total_notif': 0, 'puede_ver': lambda m: False}

    total = bd.contar_tareas_pendientes(session['usuario'])

    def puede_ver(modulo: str) -> bool:
        """
        Devuelve True si el usuario actual puede acceder al módulo indicado.
        """
        _refrescar_si_es_antigua()
        return modulo in session.get('permisos', [])

    def puede(modulo: str, accion: str = "ver") -> bool:
        """Permiso granular: True si el usuario tiene módulo:acción."""
        return _puede(modulo, accion)

    return {'total_notif': total, 'puede_ver': puede_ver, 'puede': puede}


@app.context_processor
def inject_base_template():
    """Inyecta la plantilla base adecuada dependiendo de si la petición es AJAX o estándar."""
    is_ajax = (request.headers.get('X-Requested-With') == 'XMLHttpRequest' or
               request.args.get('ajax') == '1')
    return {
        'base_template': 'base_ajax.html' if is_ajax else 'base.html',
        'is_ajax': is_ajax
    }


# =============================================================================
# DECORADOR DE SEGURIDAD
# =============================================================================

def _refrescar_si_es_antigua():
    """Si la sesión no tiene es_superadmin (sesión previa al cambio), refresca permisos."""
    if 'es_superadmin' not in session and 'usuario' in session:
        u = session['usuario']
        session['es_superadmin'] = bd.es_superadmin(u)
        session['permisos'] = bd.obtener_permisos_usuario(u)
        session['permisos_dict'] = bd.obtener_permisos_desktop(u)

def _puede(modulo: str, accion: str = "ver") -> bool:
    """Verifica si el usuario en sesión tiene permiso módulo:acción."""
    _refrescar_si_es_antigua()
    return session.get('permisos_dict', {}).get(modulo, {}).get(accion, False)


def login_requerido(f):
    """
    Decorador que protege rutas privadas.

    Si el usuario no ha iniciado sesión (no tiene 'usuario' en la sesión),
    lo redirige al login mostrando un mensaje de error. Se aplica con
    @login_requerido antes de cualquier función de ruta que requiera autenticación.
    """
    @wraps(f)
    def decorador(*args, **kwargs):
        if 'usuario' not in session:
            flash(' Acceso denegado. Por favor inicia sesión.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorador


# =============================================================================
# RUTAS PÚBLICAS — No requieren autenticación
# =============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Pantalla de inicio de sesión.

    GET:  Muestra el formulario de login.
    POST: Valida las credenciales. Si son correctas, guarda la sesión y
          redirige al Dashboard. Si son incorrectas, muestra un mensaje de error.
    """
    if request.method == 'POST':
        # Se mantiene el username tal como fue escrito; la comparación
        # insensible a mayúsculas se hace en la consulta SQL con LOWER()
        username = request.form['username'].strip()
        password = request.form['password']

        # Verificar si está bloqueado antes de intentar login
        if bd.usuario_esta_bloqueado(username):
            flash(' Cuenta bloqueada por demasiados intentos fallidos. Contacta al administrador.', 'error')
            return render_template('login.html')

        datos_usuario = bd.verificar_login(username, password)

        if datos_usuario:
            # Guardar datos del usuario en la cookie de sesión (cifrada)
            session['usuario'] = datos_usuario[0]
            session['rol'] = datos_usuario[1]
            # Cargar permisos del usuario en la sesión para el sidebar dinámico
            session['permisos'] = bd.obtener_permisos_usuario(datos_usuario[0])
            session['permisos_dict'] = bd.obtener_permisos_desktop(datos_usuario[0])
            session['es_superadmin'] = bd.es_superadmin(datos_usuario[0])
            flash(f'¡Bienvenido al sistema, {datos_usuario[0].capitalize()}!', 'exito')
            bd.registrar_accion_auditoria(datos_usuario[0], 'Inicio Sesión', 'Inició sesión en el portal web')
            return redirect(url_for('inicio'))
        else:
            intentos = bd.obtener_intentos_fallidos(username)
            restantes = max(0, bd.MAX_INTENTOS - intentos)
            if restantes > 0:
                flash(f' Usuario o contraseña incorrectos. Te quedan {restantes} intento(s).', 'error')
            else:
                flash(' Cuenta bloqueada por demasiados intentos fallidos. Contacta al administrador.', 'error')

    return render_template('login.html')


# =============================================================================
# RECUPERACIÓN DE CONTRASEÑA
# =============================================================================

@app.route('/recuperar', methods=['GET', 'POST'])
def recuperar():
    """
    Paso 1: Solicitar recuperación de contraseña.
    GET:  Muestra formulario para ingresar usuario.
    POST: Busca el usuario, si tiene email, envía código de recuperación.
    """
    if request.method == 'POST':
        identificador = request.form.get('username', '').strip()
        if not identificador:
            flash(' Ingresa tu usuario o correo electrónico.', 'error')
            return render_template('recuperar.html', paso=1)

        username, email = bd.obtener_datos_recuperacion(identificador)
        if not username or not email:
            flash(' No se encontró un email asociado a esa cuenta. Contacta al administrador.', 'error')
            return render_template('recuperar.html', paso=1)

        # Generar código de 6 dígitos
        import random
        codigo = str(random.randint(100000, 999999))

        if bd.guardar_codigo_recuperacion(username, codigo):
            exito, msg = bd.enviar_correo_recuperacion(email, codigo)
            if exito:
                flash(f' Código enviado a {email}. Revisa tu bandeja de entrada.', 'exito')
                return render_template('recuperar.html', paso=2, username=username)
            else:
                flash(f' {msg}', 'error')
        else:
            flash(' Error al generar el código. Intenta de nuevo.', 'error')

    return render_template('recuperar.html', paso=1)


@app.route('/recuperar/verificar', methods=['POST'])
def recuperar_verificar():
    """
    Paso 2: Verificar el código de recuperación.
    POST: Valida el código y permite cambiar la contraseña.
    """
    username = request.form.get('username', '').strip()
    codigo = request.form.get('codigo', '').strip()

    if not username or not codigo:
        flash(' Datos incompletos.', 'error')
        return redirect(url_for('recuperar'))

    if bd.verificar_codigo_recuperacion(username, codigo):
        return render_template('recuperar.html', paso=3, username=username, codigo=codigo)
    else:
        flash(' Código inválido o expirado. Inténtalo de nuevo.', 'error')
        return render_template('recuperar.html', paso=2, username=username)


@app.route('/recuperar/cambiar', methods=['POST'])
def recuperar_cambiar():
    """
    Paso 3: Cambiar la contraseña con el código verificado.
    """
    username = request.form.get('username', '').strip()
    codigo = request.form.get('codigo', '').strip()
    password = request.form.get('password', '').strip()
    confirmar = request.form.get('confirmar', '').strip()

    if not username or not codigo or not password:
        flash(' Datos incompletos.', 'error')
        return redirect(url_for('recuperar'))

    if password != confirmar:
        flash(' Las contraseñas no coinciden.', 'error')
        return render_template('recuperar.html', paso=3, username=username, codigo=codigo)

    if not validar_complejidad_password(password):
        flash(' La contraseña debe tener al menos 8 caracteres, una mayúscula, una minúscula, un número y un carácter especial.', 'error')
        return render_template('recuperar.html', paso=3, username=username, codigo=codigo)

    if bd.cambiar_password_con_codigo(username, codigo, password):
        flash(' Contraseña actualizada correctamente. Ahora puedes iniciar sesión.', 'exito')
        return redirect(url_for('login'))
    else:
        flash(' Error al cambiar la contraseña. El código puede haber expirado.', 'error')
        return redirect(url_for('recuperar'))


# =============================================================================
# CONFIGURACIÓN DE CORREO SMTP (solo Admin)
# =============================================================================

@app.route('/admin/config_correo', methods=['GET', 'POST'])
@login_requerido
def admin_config_correo():
    """
    Página para configurar el servidor SMTP para envío de correos.
    Requiere permiso usuarios:gestionar.
    """
    if not _puede("usuarios", "gestionar"):
        flash(' No tienes permiso para configurar el correo.', 'error')
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        servidor = request.form.get('servidor', '').strip()
        puerto = request.form.get('puerto', 587, type=int)
        usuario = request.form.get('usuario', '').strip()
        password = request.form.get('password', '').strip()
        usar_tls = request.form.get('usar_tls', 'true') == 'true'
        correo_origen = request.form.get('correo_origen', '').strip()
        nombre_origen = request.form.get('nombre_origen', 'Importadora Uziel').strip()

        if not servidor or not usuario or not correo_origen:
            flash(' Los campos servidor, usuario y correo origen son obligatorios.', 'error')
        else:
            ok = bd.guardar_config_correo(servidor, puerto, usuario, password, usar_tls, correo_origen, nombre_origen)
            if ok:
                flash(' Configuración de correo guardada correctamente.', 'success')
            else:
                flash(' Error al guardar la configuración.', 'error')

    config = bd.obtener_config_correo()
    return render_template('admin_config_correo.html', config=config)


# =============================================================================

@app.route('/logout')
def logout():
    """
    Cierra la sesión del usuario eliminando todos los datos de la cookie.
    Redirige siempre al login después de cerrar sesión.
    """
    session.clear()
    flash('Has cerrado sesión correctamente.', 'exito')
    return redirect(url_for('login'))


# =============================================================================
# RUTAS PROTEGIDAS — Requieren haber iniciado sesión (@login_requerido)
# =============================================================================

@app.route('/')
@login_requerido
def inicio():
    """
    Dashboard principal del portal.

    Muestra las tarjetas de estadísticas (total productos, total clientes,
    total tareas activas) y la tabla de los últimos productos registrados.
    También pasa al template:
      - Las tareas pendientes del usuario actual (para el modal de notificación)
      - La lista de clientes (para el formulario de asignación de tareas)
      - La lista de usuarios (para el dropdown de asignación)
    """
    # Datos de inventario y estadísticas
    inventario  = bd.obtener_productos()[:10]
    total_prod  = bd.contar_productos()
    total_cli   = bd.contar_clientes()

    # Datos para el módulo de tareas en el dashboard
    lista_clientes = bd.obtener_clientes()      # dropdown "seleccionar cliente"
    lista_usuarios = bd.obtener_usuarios()       # dropdown "asignar a"
    mis_tareas     = bd.obtener_tareas_asignadas(session['usuario'])  # notificaciones

    return render_template(
        'index.html',
        productos=inventario,
        total_prod=total_prod,
        total_cli=total_cli,
        lista_clientes=lista_clientes,
        lista_usuarios=lista_usuarios,
        mis_tareas=mis_tareas
    )


@app.route('/catalogo')
@login_requerido
def catalogo():
    """
    Módulo PIM — Vista completa del catálogo de productos.
    Permite seleccionar productos para generar un PDF y, si es Admin,
    editar o eliminar registros. Incluye la foto principal de cada producto
    para mostrar miniaturas en la tabla.
    """
    inventario = bd.obtener_productos()
    return render_template('catalogo.html', productos=inventario)


@app.route('/galeria')
@login_requerido
def galeria():
    """
    Galería visual de productos con fotos.
    Sin búsqueda: muestra los últimos 10 productos con foto.
    Con búsqueda: muestra todos los resultados coincidentes sin límite.
    """
    query = request.args.get('q', '').strip()
    if query:
        # Búsqueda activa: devolver TODOS los resultados que coincidan
        resultados = bd.buscar_banco_completo(query, limite=1000)
    else:
        # Vista inicial: solo los 10 primeros
        resultados = bd.obtener_banco_completo()[:10]

    productos_galeria = []
    for r in resultados:
        # r = (sku, nombre, ruta_principal, total_fotos, id_activo)
        productos_galeria.append({
            'sku': r[0],
            'nombre': r[1],
            'ruta': r[2],
            'total_fotos': r[3],
            'id_activo': r[4] if len(r) > 4 else None
        })

    return render_template(
        'galeria.html',
        productos=productos_galeria,
        query=query,
        total=len(productos_galeria)
    )


@app.route('/api/galeria/buscar')
@login_requerido
def api_galeria_buscar():
    """API AJAX para búsqueda en galería — retorna JSON con TODOS los resultados."""
    query = request.args.get('q', '').strip()
    if query:
        bd.registrar_accion_auditoria(session.get('usuario'), 'Búsqueda en Galería', f"Buscó: '{query}'")
    if not query:
        # Si la consulta es vacía, retornamos los últimos 10 de la vista inicial por defecto
        resultados = bd.obtener_banco_completo()[:10]
        es_reciente = True
    else:
        # Sin límite artificial: devolver todos los productos que coincidan
        resultados = bd.buscar_banco_completo(query, limite=1000)
        es_reciente = False

    items = []
    for r in resultados:
        items.append({
            'sku': r[0],
            'nombre': r[1],
            'total_fotos': r[3],
            'id_activo': r[4] if len(r) > 4 else None,
            'url': url_for('detalle_producto', sku=r[0])
        })
    return {'results': items, 'total': len(items), 'es_reciente': es_reciente}


@app.route('/producto/<sku>')
@login_requerido
def detalle_producto(sku):
    """
    Ficha de detalle de un producto con galería de fotos por ángulo.

    Muestra todos los datos técnicos del producto y las fotografías
    vinculadas desde el módulo DAM, organizadas en tabs por tipo de ángulo
    (Frontal, Lateral, Detalle, En-contexto).

    Args:
        sku (str): Código SKU del producto a mostrar.
    """
    producto = bd.obtener_producto(sku)
    if not producto:
        flash(f' Producto "{sku}" no encontrado en el catálogo.', 'error')
        return redirect(url_for('catalogo'))

    activos = bd.obtener_activos_por_sku(sku)

    # Agrupar activos por ángulo para los tabs de la galería
    galeria = {}
    for activo in activos:
        angulo = activo[3]
        if angulo not in galeria:
            galeria[angulo] = []
        # Extraer ruta relativa a almacen_activos/ para servirla por el navegador
        ruta_relativa = _extraer_ruta_relativa(activo[1])
        galeria[angulo].append({
            'id': activo[0],
            'ruta': activo[1],
            'tipo': activo[2],
            'nombre': ruta_relativa or os.path.basename(activo[1].replace('\\', '/'))
        })

    total_fotos = sum(len(lista) for lista in galeria.values())

    return render_template(
        'producto_detalle.html',
        producto=producto,
        galeria=galeria,
        total_fotos=total_fotos
    )


@app.route('/activos/<path:nombre_archivo>')
def servir_activo(nombre_archivo):
    """
    Sirve archivos estáticos desde la carpeta 'almacen_activos/'.

    Esta ruta permite que el portal web acceda a las fotografías copiadas
    por la app de escritorio DAM.
    El logotipo de la empresa en Logo/ es público, el resto requiere iniciar sesión.
    """
    nombre_normalizado = nombre_archivo.replace('\\', '/')
    if not nombre_normalizado.startswith('Logo/') and 'usuario' not in session:
        return redirect(url_for('login'))

    carpeta_activos = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'almacen_activos')
    return send_from_directory(carpeta_activos, nombre_archivo)


@app.route('/subir_imagen/<sku>', methods=['POST'])
@login_requerido
def subir_imagen(sku):
    """
    Sube una imagen desde el navegador, la guarda en alta calidad como JPG
    en el filesystem y almacena una previsualización WebP comprimida en la BD.
    Solo accesible para usuarios con rol 'Admin'.
    """
    if not _puede("activos", "subir"):
        flash(' No tienes permisos para subir imágenes.', 'error')
        return redirect(url_for('detalle_producto', sku=sku))

    # Validar que el producto existe
    producto = bd.obtener_producto(sku)
    if not producto:
        flash(f' El producto con SKU "{sku}" no existe.', 'error')
        return redirect(url_for('catalogo'))

    if 'imagen' not in request.files:
        flash(' No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('detalle_producto', sku=sku))

    archivo = request.files['imagen']
    if archivo.filename == '':
        flash(' No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('detalle_producto', sku=sku))

    angulo = request.form.get('angulo', 'Principal').strip()

    # Crear carpeta del SKU
    carpeta_activos = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'almacen_activos')
    sku_safe = sku.replace('/', '-').replace('\\', '-')
    carpeta_sku = os.path.join(carpeta_activos, sku_safe)
    os.makedirs(carpeta_sku, exist_ok=True)

    # Determinar el siguiente número secuencial
    existentes = [f for f in os.listdir(carpeta_sku) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
    numeros = []
    for f in existentes:
        base = os.path.splitext(f)[0]
        try:
            numeros.append(int(base))
        except ValueError:
            pass
    siguiente = max(numeros) + 1 if numeros else 1

    nombre_jpg = f"{siguiente}.jpg"
    ruta_jpg = os.path.join(carpeta_sku, nombre_jpg)

    try:
        # Procesar imagen con Pillow
        img = Image.open(archivo)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")

        # Guardar JPG alta calidad en filesystem (mismo formato que desktop)
        if img.width > 800 or img.height > 800:
            img.thumbnail((800, 800))
        img.save(ruta_jpg, "JPEG", quality=90, optimize=True)

        # Generar WebP de baja calidad para preview en la BD
        preview = img.copy()
        if preview.width > 300 or preview.height > 300:
            preview.thumbnail((300, 300))
        buf = io.BytesIO()
        preview.save(buf, "WEBP", quality=20, optimize=True)
        preview_binary = buf.getvalue()
        buf.close()

        img.close()

        # Guardar en BD con preview
        if bd.registrar_activo_con_preview(sku, ruta_jpg, preview_binary, "Imagen", angulo):
            flash(f' Imagen subida correctamente para SKU "{sku}".', 'exito')
            bd.registrar_accion_auditoria(session.get('usuario'), 'Subida Foto', f"Subió foto {nombre_jpg} para SKU '{sku}' (Ángulo: {angulo})")
        else:
            flash(f' La imagen se guardó en disco pero no se pudo registrar en la BD.', 'error')

    except Exception as e:
        flash(f' Error al procesar la imagen: {e}', 'error')

    return redirect(url_for('detalle_producto', sku=sku))


@app.route('/api/subir_imagen/<sku>', methods=['POST'])
def api_subir_imagen(sku):
    """
    Endpoint para que la aplicación de escritorio suba imágenes.
    Usa autenticación por API Key en lugar de cookies de sesión.
    El desktop ya registró el activo en BD; aquí solo se guarda el
    archivo y se actualiza el preview_webp, evitando duplicados.

    Cabecera requerida: X-API-Key: <clave>
    Body: form-data con campo 'imagen' (archivo)
    Campo opcional: ruta_relativa (ej: 'SKU/3.jpg')
    """
    api_key = request.headers.get('X-API-Key', '')
    if api_key != API_KEY_DESKTOP:
        return {"error": "API Key inválida"}, 401

    if 'imagen' not in request.files:
        return {"error": "No se envió ninguna imagen"}, 400

    archivo = request.files['imagen']
    if archivo.filename == '':
        return {"error": "Nombre de archivo vacío"}, 400

    angulo = request.form.get('angulo', 'Principal').strip()
    ruta_relativa = request.form.get('ruta_relativa', '').strip()
    usuario = request.form.get('usuario', 'Desktop Sync').strip()

    carpeta_activos = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'almacen_activos')

    if ruta_relativa:
        # Usar la ruta exacta que envió el desktop (evita duplicados)
        nombre_jpg = os.path.basename(ruta_relativa)
        ruta_jpg = os.path.join(carpeta_activos, ruta_relativa)
        os.makedirs(os.path.dirname(ruta_jpg), exist_ok=True)
    else:
        # Fallback: calcular número secuencial (solo si no se envió ruta)
        carpeta_sku = os.path.join(carpeta_activos, sku)
        os.makedirs(carpeta_sku, exist_ok=True)
        existentes = [f for f in os.listdir(carpeta_sku) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
        numeros = []
        for f in existentes:
            base = os.path.splitext(f)[0]
            try:
                numeros.append(int(base))
            except ValueError:
                pass
        siguiente = max(numeros) + 1 if numeros else 1
        nombre_jpg = f"{siguiente}.jpg"
        ruta_jpg = os.path.join(carpeta_sku, nombre_jpg)
        ruta_relativa = os.path.join(sku, nombre_jpg)

    try:
        img = Image.open(archivo)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        if img.width > 800 or img.height > 800:
            img.thumbnail((800, 800))
        img.save(ruta_jpg, "JPEG", quality=90, optimize=True)

        preview = img.copy()
        if preview.width > 300 or preview.height > 300:
            preview.thumbnail((300, 300))
        buf = io.BytesIO()
        preview.save(buf, "WEBP", quality=20, optimize=True)
        preview_binary = buf.getvalue()
        buf.close()
        img.close()

        # Actualizar preview en el registro existente (no duplicar)
        actualizado = bd.actualizar_preview_por_ruta(ruta_relativa, preview_binary)
        if actualizado:
            bd.registrar_accion_auditoria(usuario, 'Subida Foto', f"Sincronizó foto {nombre_jpg} para SKU '{sku}' (Ángulo: {angulo})")
            return {
                "ok": True,
                "sku": sku,
                "archivo": nombre_jpg,
                "ruta": ruta_jpg,
                "accion": "preview_actualizado"
            }

        # Fallback: si no existía el registro, crearlo con ruta relativa
        if bd.registrar_activo_con_preview(sku, ruta_relativa, preview_binary, "Imagen", angulo):
            bd.registrar_accion_auditoria(usuario, 'Subida Foto', f"Sincronizó foto {nombre_jpg} para SKU '{sku}' (Ángulo: {angulo})")
            return {
                "ok": True,
                "sku": sku,
                "archivo": nombre_jpg,
                "ruta": ruta_jpg,
                "accion": "registro_creado"
            }
        else:
            return {"error": "No se pudo registrar en la BD (¿existe el SKU?)"}, 500
    except Exception as e:
        return {"error": str(e)}, 500


@app.route('/api/bulk_subir', methods=['POST'])
def api_bulk_subir():
    """
    Endpoint para subir múltiples imágenes desde la aplicación de escritorio.
    Recibe un archivo ZIP con la estructura: SKU/archivo.jpg
    Cabecera requerida: X-API-Key: <clave>
    """
    api_key = request.headers.get('X-API-Key', '')
    if api_key != API_KEY_DESKTOP:
        return {"error": "API Key inválida"}, 401

    import zipfile
    import tempfile

    if 'archivo' not in request.files:
        return {"error": "No se envió ningún archivo"}, 400
    archivo_zip = request.files['archivo']
    if archivo_zip.filename == '':
        return {"error": "Nombre de archivo vacío"}, 400

    carpeta_activos = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'almacen_activos')
    resultados = {"ok": 0, "errores": 0, "detalle": []}

    try:
        with zipfile.ZipFile(archivo_zip) as zf:
            for nombre in zf.namelist():
                # Esperamos: SKU/archivo.jpg
                nombre_normalizado = nombre.replace('\\', '/')
                partes = nombre_normalizado.split('/')
                if len(partes) < 2:
                    continue
                sku = partes[0]
                nombre_archivo = '/'.join(partes[1:])
                if not nombre_archivo or not nombre_archivo.lower().endswith(('.jpg', '.jpeg', '.png', '.webp')):
                    continue

                carpeta_sku = os.path.join(carpeta_activos, sku)
                os.makedirs(carpeta_sku, exist_ok=True)

                # Determinar número secuencial
                existentes = [f for f in os.listdir(carpeta_sku) if f.lower().endswith(('.jpg', '.jpeg', '.png', '.webp'))]
                numeros = []
                for f in existentes:
                    base = os.path.splitext(f)[0]
                    try:
                        numeros.append(int(base))
                    except ValueError:
                        pass
                siguiente = max(numeros) + 1 if numeros else 1
                nombre_jpg = f"{siguiente}.jpg"
                ruta_jpg = os.path.join(carpeta_sku, nombre_jpg)

                try:
                    data = zf.read(nombre)
                    img = Image.open(io.BytesIO(data))
                    if img.mode in ("RGBA", "P"):
                        img = img.convert("RGB")
                    if img.width > 800 or img.height > 800:
                        img.thumbnail((800, 800))
                    img.save(ruta_jpg, "JPEG", quality=90, optimize=True)

                    preview = img.copy()
                    if preview.width > 300 or preview.height > 300:
                        preview.thumbnail((300, 300))
                    buf = io.BytesIO()
                    preview.save(buf, "WEBP", quality=20, optimize=True)
                    preview_binary = buf.getvalue()
                    buf.close()
                    img.close()

                    if bd.registrar_activo_con_preview(sku, ruta_jpg, preview_binary, "Imagen", "Principal"):
                        resultados["ok"] += 1
                        resultados["detalle"].append({"sku": sku, "archivo": nombre_jpg, "estado": "ok"})
                    else:
                        resultados["errores"] += 1
                        resultados["detalle"].append({"sku": sku, "archivo": nombre_jpg, "estado": "error_bd"})
                except Exception as e:
                    resultados["errores"] += 1
                    resultados["detalle"].append({"sku": sku, "archivo": nombre_archivo, "estado": str(e)})
    except Exception as e:
        return {"error": f"Error al procesar ZIP: {e}"}, 500

    return {"ok": True, "resultados": resultados}


@app.route('/preview/<int:activo_id>')
@login_requerido
def servir_preview(activo_id):
    """
    Sirve la previsualización WebP (baja calidad) desde la base de datos.
    Si el activo no tiene preview en BD, redirige al JPG original del filesystem.
    """
    resultado = bd.obtener_preview_activo(activo_id)
    if resultado:
        preview_bytes, ruta_fallback = resultado
        if preview_bytes:
            return send_file(
                io.BytesIO(bytes(preview_bytes)),
                mimetype='image/webp'
            )
        # Fallback: extraer ruta relativa y redirigir al archivo original
        if ruta_fallback:
            rel_path = _extraer_ruta_relativa(ruta_fallback)
            if rel_path:
                return redirect(url_for('servir_activo', nombre_archivo=rel_path))

    # Si no hay nada, devolver un placeholder 1x1 transparente
    return send_file(io.BytesIO(b''), mimetype='image/png')


@app.route('/sincronizar_previews/<sku>', methods=['POST'])
@login_requerido
def sincronizar_previews(sku):
    """
    Escanea los archivos JPG en almacen_activos/<SKU>/ y genera los previews
    WebP en la BD para todos los activos que aún no tengan preview.
    Solo Admin.

    CORREGIDO: Resuelve la ruta relativa a almacen_activos/ para que funcione
    tanto con rutas absolutas (Windows/Linux) como con rutas relativas.
    """
    if not _puede("activos", "subir"):
        flash(' No tienes permisos para sincronizar.', 'error')
        return redirect(url_for('detalle_producto', sku=sku))

    activos_sin_preview = bd.obtener_activos_sin_preview(sku)
    if not activos_sin_preview:
        flash(f' Todos los activos de "{sku}" ya tienen preview.', 'info')
        return redirect(url_for('detalle_producto', sku=sku))

    carpeta_activos = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'almacen_activos')
    contador = 0
    omitidos = 0
    for activo_id, ruta_archivo in activos_sin_preview:
        # Extraer ruta relativa y resolver contra la carpeta real del servidor
        ruta_relativa = _extraer_ruta_relativa(ruta_archivo)
        if not ruta_relativa:
            print(f" [Sync] No se pudo extraer ruta relativa de: {ruta_archivo}")
            omitidos += 1
            continue
        ruta_real = os.path.join(carpeta_activos, ruta_relativa)
        if not os.path.exists(ruta_real):
            print(f" [Sync] Archivo no encontrado, se omitirá: {ruta_real}")
            omitidos += 1
            continue
        try:
            with Image.open(ruta_real) as img:
                if img.mode in ("RGBA", "P"):
                    img = img.convert("RGB")
                preview = img.copy()
                if preview.width > 300 or preview.height > 300:
                    preview.thumbnail((300, 300))
                buf = io.BytesIO()
                preview.save(buf, "WEBP", quality=20, optimize=True)
                preview_binary = buf.getvalue()
                buf.close()

                if bd.actualizar_preview_activo(activo_id, preview_binary):
                    contador += 1
        except Exception as e:
            print(f" [Sync] Error al procesar {ruta_real}: {e}")
            omitidos += 1

    if omitidos > 0:
        flash(f' Sincronización: {contador} preview(s) generado(s), {omitidos} omitido(s) (archivos no encontrados en el servidor).', 'exito')
    else:
        flash(f' Sincronización completada: {contador} preview(s) generado(s) para "{sku}".', 'exito')
    return redirect(url_for('detalle_producto', sku=sku))


@app.route('/clientes')
@login_requerido
def clientes():
    """
    Módulo CRM — Directorio completo de clientes activos.
    Muestra RIF, nombre de empresa, teléfono y correo de cada cliente.
    """
    lista_clientes = bd.obtener_clientes()
    return render_template('clientes.html', clientes=lista_clientes)


@app.route('/cliente/<path:rif>')
@login_requerido
def cliente_detalle(rif):
    """
    Ficha completa de un cliente: datos de contacto, tareas asociadas
    y cotizaciones generadas para él.

    Args:
        rif (str): RIF del cliente (puede contener guiones y barras).
    """
    cliente = bd.obtener_cliente(rif)
    if not cliente:
        flash(' Cliente no encontrado.', 'error')
        return redirect(url_for('clientes'))

    usuario_actual = session['usuario'].lower()
    tareas_totales = bd.obtener_tareas_por_cliente(rif)
    tareas = [
        t for t in tareas_totales
        if (t[6] or '').lower() == usuario_actual or (t[5] or '').lower() == usuario_actual
    ]

    # Las alianzas se muestran según permiso
    alianzas = []
    if _puede("cotizaciones", "ver"):
        alianzas = bd.obtener_alianzas_por_aliado_rif(rif)

    return render_template(
        'cliente_detalle.html',
        cliente=cliente,
        tareas=tareas,
        alianzas=alianzas,
        puede_alianzas=_puede("cotizaciones", "ver")
    )


@app.route('/cliente/<path:rif>/editar', methods=['POST'])
@login_requerido
def cliente_editar(rif):
    """
    Actualiza los datos de contacto y ubicación de un cliente.
    Solo accesible para usuarios con rol 'Admin'.

    Args:
        rif (str): RIF del cliente a actualizar.
    """
    if not _puede("clientes", "editar"):
        flash(' No tienes permisos para editar clientes.', 'error')
        return redirect(url_for('cliente_detalle', rif=rif))

    nombre_empresa = request.form.get('nombre_empresa', '').strip()
    telefono       = request.form.get('telefono', '').strip()
    correo         = request.form.get('correo', '').strip()
    direccion      = request.form.get('direccion', '').strip()
    pais           = request.form.get('pais', '').strip()
    estado         = request.form.get('estado', '').strip()
    municipio      = request.form.get('municipio', '').strip()

    if not nombre_empresa:
        flash(' El nombre de la empresa es obligatorio.', 'error')
        return redirect(url_for('cliente_detalle', rif=rif))

    ok = bd.actualizar_cliente(rif, nombre_empresa, telefono, correo, direccion, pais, estado, municipio)
    if ok:
        flash(' Datos del cliente actualizados correctamente.', 'success')
        bd.registrar_accion_auditoria(session.get('usuario'), 'Edición Cliente', f"Editó cliente '{nombre_empresa}' (RIF: {rif})")
    else:
        flash(' Error al actualizar el cliente. Intenta de nuevo.', 'error')

    return redirect(url_for('cliente_detalle', rif=rif))


@app.route('/cliente/<path:rif>/eliminar', methods=['POST'])
@login_requerido
def cliente_eliminar(rif):
    """
    Elimina un cliente del CRM.
    Solo accesible para usuarios con el permiso 'clientes:eliminar'.
    """
    if not _puede("clientes", "eliminar"):
        flash(' No tienes permisos para eliminar clientes.', 'error')
        return redirect(url_for('cliente_detalle', rif=rif))

    cliente = bd.obtener_cliente(rif)
    if not cliente:
        flash(' Cliente no encontrado.', 'error')
        return redirect(url_for('clientes'))

    nombre_empresa = cliente[1]
    ok = bd.eliminar_cliente(rif)
    if ok:
        flash(f' Cliente "{nombre_empresa}" eliminado correctamente.', 'exito')
        bd.registrar_accion_auditoria(session.get('usuario'), 'Eliminar Cliente', f"Eliminó cliente '{nombre_empresa}' (RIF: {rif})")
        return redirect(url_for('clientes'))
    else:
        flash(' Error al eliminar el cliente de la base de datos.', 'error')
        return redirect(url_for('cliente_detalle', rif=rif))


@app.route('/nuevo_producto', methods=['GET', 'POST'])
@login_requerido
def nuevo_producto():
    """
    Formulario para registrar un nuevo producto en el inventario.
    Solo accesible para usuarios con rol 'Admin'.

    GET:  Muestra el formulario vacío.
    POST: Valida y guarda el nuevo producto. Si el SKU ya existe,
          muestra un error sin perder los datos del formulario.
    """
    if not _puede("productos", "agregar"):
        flash(' No tienes permisos para agregar productos.', 'error')
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        sku = request.form['sku'].strip().upper()
        nombre = request.form['nombre'].strip()
        descripcion = request.form['descripcion'].strip()
        marca = request.form['marca'].strip()
        compatibilidad = request.form['compatibilidad'].strip()
        precio = request.form['precio']

        if bd.registrar_producto(sku, nombre, descripcion, marca, compatibilidad, precio):
            flash(f' ¡Producto {sku} agregado exitosamente al inventario!', 'exito')
            bd.registrar_accion_auditoria(session.get('usuario'), 'Creación Producto', f"Creó producto '{nombre}' (SKU: {sku})")
            return redirect(url_for('catalogo'))
        else:
            flash(f' Error al registrar. El SKU "{sku}" ya podría existir. Verifique.', 'error')

    return render_template('nuevo_producto.html')


@app.route('/nuevo_cliente', methods=['GET', 'POST'])
@login_requerido
def nuevo_cliente():
    """
    Formulario para registrar un nuevo cliente en el CRM.
    Solo accesible para usuarios con rol 'Admin'.

    GET:  Muestra el formulario vacío.
    POST: Valida y guarda el nuevo cliente. Si el RIF ya existe, muestra un error.
    """
    if not _puede("clientes", "agregar"):
        flash(' No tienes permisos para agregar clientes.', 'error')
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        rif = request.form['rif'].strip()
        nombre_empresa = request.form['nombre_empresa'].strip()
        telefono = request.form['telefono'].strip()
        correo = request.form['correo'].strip()
        direccion = request.form['direccion'].strip()
        pais = request.form.get('pais', '').strip()
        estado = request.form.get('estado', '').strip()
        municipio = request.form.get('municipio', '').strip()

        if bd.registrar_cliente(rif, nombre_empresa, telefono, correo, direccion, pais, estado, municipio):
            flash(f' ¡Cliente "{nombre_empresa}" registrado exitosamente!', 'exito')
            bd.registrar_accion_auditoria(session.get('usuario'), 'Creación Cliente', f"Creó cliente '{nombre_empresa}' (RIF: {rif})")
            return redirect(url_for('clientes'))
        else:
            flash(' Error al registrar. Verifique que el RIF no esté duplicado.', 'error')

    return render_template('nuevo_cliente.html')


@app.route('/editar/<sku>', methods=['GET', 'POST'])
@login_requerido
def editar_producto(sku):
    """
    Formulario para editar los datos de un producto existente.
    Solo accesible para usuarios con rol 'Admin'. El SKU no puede cambiarse.

    GET:  Carga el formulario con los datos actuales del producto.
    POST: Guarda los cambios y redirige al catálogo.
    """
    if not _puede("productos", "editar"):
        flash(' No tienes permisos para editar productos.', 'error')
        return redirect(url_for('inicio'))

    if request.method == 'POST':
        nombre = request.form['nombre'].strip()
        descripcion = request.form['descripcion'].strip()
        marca = request.form['marca'].strip()
        compatibilidad = request.form['compatibilidad'].strip()
        precio = request.form['precio']

        if bd.actualizar_producto(sku, nombre, descripcion, marca, compatibilidad, precio):
            flash(f' Producto {sku} actualizado correctamente.', 'exito')
            bd.registrar_accion_auditoria(session.get('usuario'), 'Edición Producto', f"Editó producto '{nombre}' (SKU: {sku})")
            return redirect(url_for('catalogo'))
        else:
            flash(f' No se pudo actualizar el producto {sku}.', 'error')
    else:
        # GET: cargar datos actuales del producto para pre-llenar el formulario
        producto = bd.obtener_producto(sku)
        if not producto:
            flash(f' El producto con SKU "{sku}" no fue encontrado.', 'error')
            return redirect(url_for('catalogo'))
        return render_template('editar.html', producto=producto)


@app.route('/eliminar/<sku>', methods=['POST'])
@login_requerido
def eliminar_producto(sku):
    """
    Elimina permanentemente un producto del inventario.
    Solo accesible para usuarios con rol 'Admin'.
    La confirmación visual se maneja con un onclick en el HTML.
    """
    if not _puede("productos", "eliminar"):
        flash(' No tienes permisos para eliminar productos.', 'error')
        return redirect(url_for('inicio'))

    producto = bd.obtener_producto(sku)
    nombre = producto[2] if (producto and len(producto) > 2) else "Desconocido"
    if bd.eliminar_producto(sku):
        flash(f' El producto {sku} fue eliminado del inventario.', 'exito')
        bd.registrar_accion_auditoria(session.get('usuario'), 'Eliminar Producto', f"Eliminó producto '{nombre}' (SKU: {sku})")
    else:
        flash(f' No se pudo eliminar el producto {sku}.', 'error')

    return redirect(url_for('catalogo'))


# =============================================================================
# MÓDULO DE TAREAS — Asignación y seguimiento de trabajo interno
# =============================================================================

@app.route('/tareas')
@login_requerido
def tareas():
    """
    Página de gestión de tareas del sistema.
    Cada usuario (incluyendo administradores) ve únicamente las tareas
    en las que es creador (creado_por) o responsable asignado (asignado_a).
    """
    usuario_actual = session['usuario']
    lista_tareas = bd.obtener_tareas_visibles_usuario(usuario_actual)
    
    # Cargar clientes y usuarios solo para quienes tengan permiso de gestionar
    if _puede("tareas", "gestionar"):
        lista_clientes  = bd.obtener_clientes()
        lista_usuarios  = bd.obtener_usuarios()
    else:
        lista_clientes  = []
        lista_usuarios  = []

    return render_template(
        'tareas.html',
        lista_tareas=lista_tareas,
        lista_clientes=lista_clientes,
        lista_usuarios=lista_usuarios
    )


@app.route('/asignar_tarea', methods=['POST'])
@login_requerido
def asignar_tarea():
    """
    Crea una nueva tarea a partir del formulario del dashboard o de la
    página de tareas. Solo accesible para usuarios con rol 'Admin'.

    Datos esperados del formulario (POST):
        cliente_rif    — RIF del cliente seleccionado en el dropdown
        cliente_nombre — Nombre de la empresa (campo oculto del dropdown)
        asignado_a     — Username del responsable
        tipo_tarea     — Categoría del trabajo
        descripcion    — Descripción libre
        fecha_limite   — Fecha tope (YYYY-MM-DD)
    """
    if not _puede("tareas", "gestionar"):
        flash(' No tienes permisos para asignar tareas.', 'error')
        return redirect(url_for('inicio'))

    # Leer y limpiar campos del formulario
    cliente_rif    = request.form.get('cliente_rif', '').strip()
    cliente_nombre = request.form.get('cliente_nombre', '').strip()
    asignado_a     = request.form.get('asignado_a', '').strip()
    tipo_tarea     = request.form.get('tipo_tarea', '').strip()
    descripcion    = request.form.get('descripcion', '').strip()
    fecha_limite   = request.form.get('fecha_limite', '').strip()
    creado_por     = session['usuario']

    # Validar que todos los campos obligatorios estén completos
    if not all([cliente_rif, asignado_a, tipo_tarea, fecha_limite]):
        flash(' Completa todos los campos obligatorios de la tarea.', 'error')
        return redirect(request.referrer or url_for('tareas'))

    # Guardar la tarea en la base de datos
    if bd.crear_tarea(cliente_rif, cliente_nombre, asignado_a,
                      tipo_tarea, descripcion, fecha_limite, creado_por):
        flash(f' Tarea asignada a "{asignado_a}" correctamente.', 'exito')
    else:
        flash(' No se pudo crear la tarea. Intenta de nuevo.', 'error')

    # Regresar a la página desde donde se envió el formulario
    return redirect(request.referrer or url_for('tareas'))


@app.route('/actualizar_tarea/<int:tarea_id>', methods=['POST'])
@login_requerido
def actualizar_tarea(tarea_id):
    """
    Actualiza el estado de una tarea (Pendiente → En Progreso → Completada).
    Solo el creador o la persona asignada pueden actualizar el estado.
    """
    tarea = bd.obtener_tarea_por_id(tarea_id)
    if not tarea:
        flash(' Tarea no encontrada.', 'error')
        return redirect(url_for('tareas'))

    usuario_actual = session['usuario'].lower()
    creado_por = (tarea[8] or '').lower()
    asignado_a = (tarea[3] or '').lower()

    if usuario_actual != creado_por and usuario_actual != asignado_a:
        flash(' No tienes permisos para actualizar esta tarea.', 'error')
        return redirect(url_for('tareas'))

    nuevo_estado = request.form.get('nuevo_estado', '').strip()

    # Validar que el estado recibido sea uno de los permitidos
    estados_validos = {'Pendiente', 'En Progreso', 'Completada'}
    if nuevo_estado not in estados_validos:
        flash(' Estado no válido.', 'error')
        return redirect(url_for('tareas'))

    if bd.actualizar_estado_tarea(tarea_id, nuevo_estado):
        flash(f' Tarea #{tarea_id} marcada como "{nuevo_estado}".', 'exito')
    else:
        flash(f' No se pudo actualizar la tarea #{tarea_id}.', 'error')

    return redirect(url_for('tareas'))


@app.route('/tarea/<int:tarea_id>/eliminar', methods=['POST'])
@login_requerido
def eliminar_tarea(tarea_id):
    """
    Elimina una tarea del tablero.
    Solo accesible si el usuario tiene permiso 'tareas:gestionar' Y es el creador de la tarea.
    """
    if not _puede("tareas", "gestionar"):
        flash(' No tienes permisos para eliminar tareas.', 'error')
        return redirect(url_for('tareas'))

    tarea = bd.obtener_tarea_por_id(tarea_id)
    if not tarea:
        flash(' Tarea no encontrada.', 'error')
        return redirect(url_for('tareas'))

    usuario_actual = session['usuario'].lower()
    creado_por = (tarea[8] or '').lower()

    if usuario_actual != creado_por:
        flash(' Solo el creador de la tarea puede eliminarla.', 'error')
        return redirect(url_for('tareas'))

    if bd.eliminar_tarea(tarea_id):
        flash(' Tarea eliminada exitosamente.', 'exito')
    else:
        flash(' Error al eliminar la tarea.', 'error')

    return redirect(url_for('tareas'))


# =============================================================================
# GENERADOR DE PDF
# =============================================================================

@app.route('/generar_pdf', methods=['POST'])
@login_requerido
def generar_pdf():
    skus_seleccionados = request.form.getlist('skus_seleccionados')

    if not skus_seleccionados:
        flash(' Selecciona al menos un producto (casilla PDF) antes de generar.', 'error')
        return redirect(url_for('catalogo'))

    # Colores corporativos (RGB 0.0-1.0) — idénticos al generador de escritorio
    AZUL_EMP   = (0.18, 0.27, 0.86)
    NEGRO      = (0.17, 0.22, 0.31)
    VERDE_PREC = (0.15, 0.68, 0.37)
    GRIS       = (0.55, 0.60, 0.68)
    LINEA      = (0.87, 0.88, 0.93)

    MARGEN_X = 50
    ANCHO_PAG, ALTO_PAG = letter
    ALTO_FILA = 72

    buffer = io.BytesIO()
    c = pdf_canvas.Canvas(buffer, pagesize=letter)
    from datetime import datetime

    def encabezado():
        c.setFillColorRGB(*AZUL_EMP)
        c.rect(0, ALTO_PAG - 80, ANCHO_PAG, 80, fill=True, stroke=False)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 20)
        c.drawString(MARGEN_X, ALTO_PAG - 48, "IMPORTADORA UZIEL C.A.")
        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.8, 0.88, 1.0)
        c.drawString(MARGEN_X, ALTO_PAG - 65, "Catálogo de Productos — Listado General")
        c.drawString(MARGEN_X, ALTO_PAG - 75, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    def pie():
        c.setFont("Helvetica", 8)
        c.setFillColorRGB(*GRIS)
        c.drawString(MARGEN_X, 35, "Documento generado automáticamente por el Sistema de Información Uziel.")
        c.drawRightString(ANCHO_PAG - MARGEN_X, 35, datetime.now().strftime("Generado el %d/%m/%Y a las %H:%M"))

    def dibujar_fila(sku, nombre, marca, compatibilidad, precio, ruta_img, y, idx):
        # Fondo alternado
        c.setFillColorRGB(0.97, 0.98, 1.0) if idx % 2 == 0 else c.setFillColorRGB(1, 1, 1)
        c.rect(MARGEN_X, y - ALTO_FILA, ANCHO_PAG - 2 * MARGEN_X, ALTO_FILA, fill=True, stroke=False)

        # Dibujar thumbnail
        thumb_ancho = 50
        thumb_alto = 50
        thumb_x = MARGEN_X + 10
        thumb_y = y - thumb_alto - 11

        # Resolver ruta local en el servidor
        ruta_local = None
        if ruta_img:
            ruta_rel = _extraer_ruta_relativa(ruta_img)
            carpeta_activos = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'almacen_activos')
            ruta_local = os.path.join(carpeta_activos, ruta_rel) if ruta_rel else None

        has_thumb = False
        if ruta_local and os.path.exists(ruta_local):
            try:
                # Dibujar borde y la imagen
                c.setStrokeColorRGB(*LINEA)
                c.setLineWidth(0.5)
                c.roundRect(thumb_x, thumb_y, thumb_ancho, thumb_alto, 4, fill=False, stroke=True)
                c.drawImage(ruta_local, thumb_x, thumb_y, width=thumb_ancho, height=thumb_alto, preserveAspectRatio=True, mask="auto")
                has_thumb = True
            except Exception as e:
                print(f"[PDF Gen] Error al dibujar imagen local {ruta_local}: {e}")
                pass

        # Fallback: Obtener bytes de la imagen directamente de la base de datos si no existe el archivo local
        if not has_thumb:
            try:
                preview_bytes = bd.obtener_preview_principal_por_sku(sku)
                if preview_bytes:
                    c.setStrokeColorRGB(*LINEA)
                    c.setLineWidth(0.5)
                    c.roundRect(thumb_x, thumb_y, thumb_ancho, thumb_alto, 4, fill=False, stroke=True)
                    # ReportLab permite usar un objeto tipo archivo como BytesIO para la imagen
                    c.drawImage(io.BytesIO(preview_bytes), thumb_x, thumb_y, width=thumb_ancho, height=thumb_alto, preserveAspectRatio=True, mask="auto")
                    has_thumb = True
            except Exception as e:
                print(f"[PDF Gen] Error al dibujar imagen desde la BD para {sku}: {e}")
                pass

        if not has_thumb:
            # Dibujar placeholder
            c.setStrokeColorRGB(*LINEA)
            c.setLineWidth(0.5)
            c.setFillColorRGB(0.96, 0.97, 0.99)
            c.roundRect(thumb_x, thumb_y, thumb_ancho, thumb_alto, 4, fill=True, stroke=True)
            c.setFont("Helvetica-Oblique", 7)
            c.setFillColorRGB(*GRIS)
            c.drawCentredString(thumb_x + thumb_ancho / 2, thumb_y + thumb_alto / 2 - 2, "Sin imagen")

        labels = ["SKU:", "Producto:", "Marca:", "Compatibilidad:", "Precio:"]
        valores = [str(sku), str(nombre), str(marca), str(compatibilidad), f"$ {precio}" if precio is not None else "—"]
        ancho_label = 68
        y_linea = y - 16
        x_texto = MARGEN_X + 70

        for i, (label, valor) in enumerate(zip(labels, valores)):
            c.setFont("Helvetica-Bold", 8)
            c.setFillColorRGB(*GRIS)
            c.drawString(x_texto, y_linea, label)
            c.setFont("Helvetica", 8.5)
            if i == 4:
                c.setFillColorRGB(*VERDE_PREC)
                c.setFont("Helvetica-Bold", 9)
            else:
                c.setFillColorRGB(*NEGRO)

            texto = str(valor)
            limite_ancho = ANCHO_PAG - MARGEN_X - ancho_label - MARGEN_X - 70
            if c.stringWidth(texto, "Helvetica", 8.5) > limite_ancho:
                while c.stringWidth(texto + "...", "Helvetica", 8.5) > limite_ancho:
                    texto = texto[:-1]
                texto += "..."

            c.drawString(x_texto + ancho_label, y_linea, texto)
            y_linea -= 12

        c.setStrokeColorRGB(*LINEA)
        c.setLineWidth(0.5)
        c.line(MARGEN_X, y - ALTO_FILA, ANCHO_PAG - MARGEN_X, y - ALTO_FILA)

    encabezado()
    y = ALTO_PAG - 115
    idx = 0

    for sku in skus_seleccionados:
        prod = bd.obtener_producto(sku)
        if not prod:
            continue
        if y < ALTO_FILA + 50:
            pie()
            c.showPage()
            encabezado()
            y = ALTO_PAG - 115

        # Obtener ruta de la imagen
        ruta_img = bd.obtener_activo_principal(sku)
        if not ruta_img:
            prod_img_data = bd.obtener_producto_con_imagen(sku)
            ruta_img = prod_img_data[4] if prod_img_data and len(prod_img_data) > 4 else None

        dibujar_fila(prod.sku, prod.nombre, prod.marca, prod.compatibilidad, prod.precio, ruta_img, y, idx)
        y -= ALTO_FILA
        idx += 1

    pie()
    c.save()
    buffer.seek(0)

    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{NOMBRE_ARCHIVO_PDF}"'
    return response


# =============================================================================
# MÓDULO DE REPORTES
# =============================================================================

@app.route('/reportes')
@login_requerido
def reportes():
    """Página de reportes con selector de fechas personalizado."""
    from datetime import datetime, timedelta

    hoy = datetime.now()

    # Leer fechas de query params o usar valores por defecto (mes actual)
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    termino = request.args.get('termino', '')

    if desde and hasta:
        fecha_desde = desde
        fecha_hasta = hasta
    else:
        # Por defecto: mes actual
        inicio_mes = hoy.replace(day=1)
        if hoy.month == 12:
            fin_mes = hoy.replace(year=hoy.year + 1, month=1, day=1)
        else:
            fin_mes = hoy.replace(month=hoy.month + 1, day=1)
        fecha_desde = inicio_mes.strftime("%Y-%m-%d")
        fecha_hasta = fin_mes.strftime("%Y-%m-%d")

    # Datos del reporte
    datos = bd.obtener_datos_reporte(fecha_desde, fecha_hasta, termino=termino)
    datos_productos = bd.obtener_productos_por_fecha(fecha_desde, fecha_hasta, pagina=1, por_pagina=10)
    eventos_especiales = bd.obtener_ultimos_eventos_especiales()
    logs_pag = bd.obtener_logs_auditoria_paginados(fecha_desde, fecha_hasta, pagina=1, por_pagina=5, termino=termino)

    # Semana actual para referencia rapida
    diasem = hoy.weekday()
    domingo_pasado = hoy - timedelta(days=(diasem + 1) % 7)
    domingo_siguiente = domingo_pasado + timedelta(days=7)

    return render_template(
        'reportes.html',
        desde=fecha_desde,
        hasta=fecha_hasta,
        termino=termino,
        datos=datos,
        productos_pag=datos_productos,
        logs_pag=logs_pag,
        eventos_especiales=eventos_especiales,
        semana_inicio=domingo_pasado.strftime("%d/%m/%Y"),
        semana_fin=domingo_siguiente.strftime("%d/%m/%Y")
    )


@app.route('/api/reporte_productos')
@login_requerido
def api_reporte_productos():
    """API JSON para paginacion de productos en reporte."""
    from datetime import datetime, timedelta

    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    pagina = request.args.get('pagina', 1, type=int)

    if not desde or not hasta:
        return {"productos": [], "total": 0, "html": ""}

    datos = bd.obtener_productos_por_fecha(desde, hasta, pagina=pagina, por_pagina=10)

    html = ""
    for p in datos["productos"]:
        precio = f"${float(p[4]):,.2f}" if p[4] else "—"
        html += f"""<tr>
            <td><span class="sku-badge">{p[0]}</span></td>
            <td>{p[1]}</td>
            <td>{p[2]}</td>
            <td class="precio">{precio}</td>
        </tr>"""

    return {
        "productos": datos["productos"],
        "total": datos["total"],
        "pagina": datos["pagina"],
        "total_paginas": datos["total_paginas"],
        "html": html
    }


@app.route('/api/reporte_auditoria')
@login_requerido
def api_reporte_auditoria():
    """API JSON para paginacion de la bitácora de auditoría en reporte."""
    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    pagina = request.args.get('pagina', 1, type=int)
    termino = request.args.get('termino', '')

    if not desde or not hasta:
        return {"logs": [], "total": 0, "html": ""}

    datos = bd.obtener_logs_auditoria_paginados(desde, hasta, pagina=pagina, por_pagina=5, termino=termino)

    html = ""
    for log in datos["logs"]:
        fecha_str = log[4] if isinstance(log[4], str) else log[4].strftime('%d/%m/%Y %H:%M:%S')
        html += f"""<tr>
            <td style="white-space: nowrap; font-weight: 500;">{fecha_str}</td>
            <td><strong style="color:var(--text);">{log[1]}</strong></td>
            <td><span class="sku-badge" style="background:var(--primary-light); color:var(--primary-dark);">{log[2]}</span></td>
            <td style="color:var(--text-muted); font-size:13px;">{log[3]}</td>
        </tr>"""

    return {
        "logs": datos["logs"],
        "total": datos["total"],
        "pagina": datos["pagina"],
        "total_paginas": datos["total_paginas"],
        "html": html
    }


@app.route('/reporte_pdf/<tipo>')
@app.route('/reporte_pdf')
@login_requerido
def reporte_pdf(tipo=None):
    """Descarga el PDF de reporte semanal, mensual o personalizado."""
    from datetime import datetime, timedelta

    hoy = datetime.now()

    desde = request.args.get('desde', '')
    hasta = request.args.get('hasta', '')
    termino = request.args.get('termino', '')

    if desde and hasta:
        fecha_inicio = desde
        fecha_fin = hasta
        tipo_reporte = "Personalizado"
    elif tipo == "semanal":
        diasem = hoy.weekday()
        fecha_inicio = hoy - timedelta(days=(diasem + 1) % 7)
        fecha_fin = fecha_inicio + timedelta(days=7)
        tipo_reporte = "Semanal"
    elif tipo == "mensual":
        fecha_inicio = hoy.replace(day=1)
        if hoy.month == 12:
            fecha_fin = hoy.replace(year=hoy.year + 1, month=1, day=1)
        else:
            fecha_fin = hoy.replace(month=hoy.month + 1, day=1)
        tipo_reporte = "Mensual"
    else:
        flash("Especifica un tipo de reporte o fechas.", "error")
        return redirect(url_for('reportes'))

    fi = fecha_inicio if isinstance(fecha_inicio, str) else fecha_inicio.strftime("%Y-%m-%d")
    ff = fecha_fin if isinstance(fecha_fin, str) else fecha_fin.strftime("%Y-%m-%d")

    datos = bd.obtener_datos_reporte(fi, ff, termino=termino)
    datos["fecha_inicio"] = fi
    datos["fecha_fin"] = ff

    prod_data = bd.obtener_productos_por_fecha(fi, ff, pagina=1, por_pagina=500)
    productos_list = prod_data.get("productos", [])

    buffer = generar_reporte_pdf(datos, tipo_reporte, productos_list=productos_list)

    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="Reporte_{tipo_reporte}_{hoy.strftime("%Y%m%d")}.pdf"'
    return response


# =============================================================================
# MÓDULO ALIANZAS COMERCIALES (GAAE)
# =============================================================================

@app.route('/alianzas')
@login_requerido
def alianzas():
    """
    Lista todas las alianzas registradas.
    Permite filtrar por estado mediante el parámetro GET ?estado=...
    """
    if not _puede("cotizaciones", "ver"):
        flash(' No tienes permisos para acceder a alianzas.', 'error')
        return redirect(url_for('inicio'))
    estado_filtro = request.args.get('estado', '')
    lista = bd.obtener_alianzas(estado_filtro if estado_filtro else None)
    estados = ['Borrador', 'Autorizada', 'Entregada', 'Incumplida', 'Completada']
    clientes_lista = bd.obtener_clientes()
    productos_lista = bd.obtener_productos()
    return render_template(
        'alianzas.html',
        alianzas=lista,
        estados=estados,
        estado_activo=estado_filtro,
        clientes=clientes_lista,
        productos=productos_lista
    )


@app.route('/alianza/nueva', methods=['GET', 'POST'])
@login_requerido
def alianza_nueva():
    """
    Crear una nueva alianza.

    GET:  Muestra el formulario con selector de aliado y buscador de activos.
    POST: Valida los datos, guarda la alianza y redirige al detalle.
    """
    if not _puede("cotizaciones", "crear"):
        flash(' No tienes permisos para registrar alianzas.', 'error')
        return redirect(url_for('alianzas'))

    if request.method == 'POST':
        cliente_rif    = request.form.get('cliente_rif', '').strip()
        cliente_nombre = request.form.get('cliente_nombre', '').strip()
        notas          = request.form.get('notas', '').strip()

        # Recopilar ítems del formulario (arrays paralelos)
        skus       = request.form.getlist('sku[]')
        nombres    = request.form.getlist('nombre[]')
        cantidades = request.form.getlist('cantidad[]')
        precios    = request.form.getlist('precio_unitario[]')

        if not cliente_rif or not cliente_nombre or not skus:
            flash(' Debes seleccionar un aliado y agregar al menos un activo.', 'error')
        else:
            # Parseo completo primero — si algo falla, no se guarda nada
            items = []
            error_validacion = False
            for i, sku in enumerate(skus):
                if not sku:
                    continue
                try:
                    cant  = int(cantidades[i])
                    prec  = float(precios[i])
                    if cant <= 0 or prec < 0:
                        raise ValueError("Valores fuera de rango")
                except (ValueError, IndexError):
                    flash(' Cantidad o valor unitario inválido.', 'error')
                    error_validacion = True
                    break
                items.append({
                    'sku': sku,
                    'nombre': nombres[i] if i < len(nombres) else sku,
                    'cantidad': cant,
                    'precio_unitario': prec
                })

            # Solo persistir si NO hubo error y hay al menos un ítem válido
            if not error_validacion and items:
                cot_id = bd.crear_alianza(
                    cliente_rif, session['usuario'], items, notas
                )
                if cot_id:
                    flash(' Alianza comercial registrada exitosamente.', 'success')
                    return redirect(url_for('alianza_detalle', alianza_id=cot_id))
                else:
                    flash(' Error al guardar la alianza. Intenta de nuevo.', 'error')
            elif not error_validacion:
                flash(' Debes agregar al menos un activo válido.', 'error')

    return redirect(url_for('alianzas'))


@app.route('/alianza/<int:alianza_id>')
@login_requerido
def alianza_detalle(alianza_id):
    """
    Muestra la alianza completa con todos sus ítems, estado y opciones de acción.
    """
    if not _puede("cotizaciones", "ver"):
        flash(' No tienes permisos para ver alianzas.', 'error')
        return redirect(url_for('inicio'))
    datos = bd.obtener_alianza_con_items(alianza_id)
    if not datos:
        flash(' Alianza comercial no encontrada.', 'error')
        return redirect(url_for('alianzas'))
    estados = ['Borrador', 'Autorizada', 'Entregada', 'Incumplida', 'Completada']
    return render_template(
        'alianza_detalle.html',
        cab=datos['cabecera'],
        items=datos['items'],
        auditoria=datos['auditoria'],
        estados=estados
    )


@app.route('/alianza/<int:alianza_id>/editar', methods=['GET', 'POST'])
@login_requerido
def alianza_editar(alianza_id):
    if not _puede("cotizaciones", "crear"):
        flash(' No tienes permisos para editar alianzas.', 'error')
        return redirect(url_for('alianza_detalle', alianza_id=alianza_id))
    
    if request.method == 'POST':
        notes = request.form.get('notas', '').strip()
        skus = request.form.getlist('sku[]')
        nombres = request.form.getlist('nombre[]')
        cantidades = request.form.getlist('cantidad[]')
        precios = request.form.getlist('precio_unitario[]')
        
        if not skus:
            flash(' Debes agregar al menos un activo.', 'error')
        else:
            items = []
            error_validacion = False
            for i, sku in enumerate(skus):
                if not sku:
                    continue
                try:
                    cant = int(cantidades[i])
                    prec = float(precios[i])
                    if cant <= 0 or prec < 0:
                        raise ValueError("Valores fuera de rango")
                except (ValueError, IndexError):
                    flash(' Cantidad o valor unitario inválido.', 'error')
                    error_validacion = True
                    break
                items.append({
                    'sku': sku,
                    'nombre': nombres[i] if i < len(nombres) else sku,
                    'cantidad': cant,
                    'precio_unitario': prec
                })
            
            if not error_validacion and items:
                ok = bd.actualizar_alianza(alianza_id, items, notes)
                if ok:
                    flash(' Alianza comercial actualizada correctamente.', 'success')
                    return redirect(url_for('alianza_detalle', alianza_id=alianza_id))
                else:
                    flash(' Error al actualizar la alianza.', 'error')
                    
    datos = bd.obtener_alianza_con_items(alianza_id)
    if not datos:
        flash(' Alianza comercial no encontrada.', 'error')
        return redirect(url_for('alianzas'))
        
    productos_lista = bd.obtener_productos()
    return render_template(
        'alianza_editar.html',
        cab=datos['cabecera'],
        items=datos['items'],
        productos=productos_lista
    )


@app.route('/alianza/<int:alianza_id>/estado', methods=['POST'])
@login_requerido
def alianza_estado(alianza_id):
    """
    Cambia el estado de una alianza.
    POST param: nuevo_estado
    """
    if not _puede("cotizaciones", "crear"):
        flash(' No tienes permisos para cambiar el estado.', 'error')
        return redirect(url_for('alianza_detalle', alianza_id=alianza_id))

    nuevo_estado = request.form.get('nuevo_estado', '').strip()
    estados_validos = ['Borrador', 'Autorizada', 'Entregada', 'Incumplida', 'Completada']
    if nuevo_estado not in estados_validos:
        flash(' Estado no válido.', 'error')
    else:
        ok = bd.actualizar_estado_alianza(alianza_id, nuevo_estado)
        if ok:
            flash(f' Estado actualizado a "{nuevo_estado}".', 'success')
        else:
            flash(' Error al actualizar el estado.', 'error')

    return redirect(url_for('alianza_detalle', alianza_id=alianza_id))


@app.route('/alianza/<int:alianza_id>/pdf')
@login_requerido
def alianza_pdf(alianza_id):
    """
    Genera y descarga el PDF de la alianza indicada.
    """
    if not _puede("cotizaciones", "ver"):
        flash(' No tienes permisos para descargar la alianza.', 'error')
        return redirect(url_for('inicio'))
    datos = bd.obtener_alianza_con_items(alianza_id)
    if not datos:
        flash(' Alianza comercial no encontrada.', 'error')
        return redirect(url_for('alianzas'))

    buffer = generar_pdf_alianza(datos)
    numero = datos['cabecera'][1]
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="Alianza_{numero}.pdf"'
    return response


@app.route('/alianza/<int:alianza_id>/eliminar', methods=['POST'])
@login_requerido
def eliminar_alianza(alianza_id):
    """
    Elimina una alianza comercial.
    Solo accesible para usuarios con permiso cotizaciones:crear.
    """
    if not _puede("cotizaciones", "crear"):
        flash(' No tienes permisos para eliminar alianzas.', 'error')
        return redirect(url_for('alianza_detalle', alianza_id=alianza_id))

    if bd.eliminar_alianza(alianza_id):
        flash(' Alianza comercial eliminada exitosamente.', 'exito')
        return redirect(url_for('alianzas'))
    else:
        flash(' Error al eliminar la alianza comercial.', 'error')
        return redirect(url_for('alianza_detalle', alianza_id=alianza_id))


# =============================================================================
# MÓDULO: GESTIÓN DE USUARIOS (solo Admin)
# =============================================================================

@app.route('/admin/usuarios')
@login_requerido
def admin_usuarios():
    """
    Panel de administración de usuarios.
    Muestra la lista completa de usuarios con sus roles y permisos.
    Requiere permiso usuarios:gestionar.
    """
    if not _puede("usuarios", "gestionar"):
        flash(' No tienes permiso para gestionar usuarios.', 'error')
        return redirect(url_for('inicio'))

    usuarios = bd.obtener_todos_usuarios()
    return render_template('admin_usuarios.html', usuarios=usuarios, modulos_acciones=bd.MODULOS_ACCIONES)


@app.route('/admin/usuario/nuevo', methods=['POST'])
@login_requerido
def admin_usuario_nuevo():
    """
    Crea un nuevo usuario en el sistema.
    Requiere permiso usuarios:gestionar.
    """
    if not _puede("usuarios", "gestionar"):
        flash(' No tienes permiso para crear usuarios.', 'error')
        return redirect(url_for('inicio'))

    username   = request.form.get('username', '').strip()
    password   = request.form.get('password', '').strip()
    rol        = request.form.get('rol', 'Empleado').strip()
    email      = request.form.get('email', '').strip()
    permisos   = request.form.get('permisos', '').strip()
    if not permisos:
        permisos = ','.join(request.form.getlist('permisos'))

    if not username or not password:
        flash(' El usuario y la contraseña son obligatorios.', 'error')
        return redirect(url_for('admin_usuarios'))

    if not validar_complejidad_password(password):
        flash(' La contraseña no cumple con los requisitos de seguridad (mínimo 8 caracteres, mayúsculas, minúsculas, números y caracteres especiales).', 'error')
        return redirect(url_for('admin_usuarios'))

    bloqueado = False
    if session.get('es_superadmin'):
        bloqueado = request.form.get('bloqueado') in ('true', 'on')
    
    # Solo un superadmin existente puede crear otro Admin
    if rol == 'Admin' and not session.get('es_superadmin'):
        rol = 'Supervisor'
    
    superadmin = (rol == 'Admin')

    ok = bd.crear_usuario(username, password, rol, permisos, email, superadmin, bloqueado)
    if ok:
        flash(f' Usuario "{username}" creado correctamente.', 'success')
        bd.registrar_accion_auditoria(session.get('usuario'), 'Creación Usuario', f"Creó usuario '{username}' (Rol: {rol})")
    else:
        flash(f' No se pudo crear el usuario "{username}". Es posible que ya exista.', 'error')

    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuario/<username>/editar', methods=['POST'])
@login_requerido
def admin_usuario_editar(username):
    """
    Actualiza nombre de usuario, rol, permisos y email.
    Requiere permiso usuarios:gestionar.
    """
    if not _puede("usuarios", "gestionar"):
        flash(' No tienes permiso para editar usuarios.', 'error')
        return redirect(url_for('inicio'))

    # (Se eliminó la restricción de que solo un superadmin puede modificar a otro superadmin, ahora cualquier Admin puede hacerlo)

    nuevo_username = request.form.get('nuevo_username', '').strip()
    nuevo_rol      = request.form.get('rol', 'Empleado').strip()
    email          = request.form.get('email', '').strip()
    permisos       = request.form.get('permisos', '').strip()
    if not permisos:
        permisos = ','.join(request.form.getlist('permisos'))

    if not nuevo_username:
        flash(' El nombre de usuario no puede quedar vacío.', 'error')
        return redirect(url_for('admin_usuarios'))

    bloqueado = False
    if session.get('es_superadmin') or session.get('rol') == 'Admin':
        bloqueado = request.form.get('bloqueado') in ('true', 'on')
        superadmin = (nuevo_rol == 'Admin')
    else:
        if nuevo_rol == 'Admin':
            nuevo_rol = 'Supervisor'  # Prevenir escalada de privilegios
        # Recuperar valores existentes para no sobreescribirlos
        datos_ex = bd.obtener_datos_completos_usuario(username)
        superadmin = bd.es_superadmin(username)
        if datos_ex:
            bloqueado = bool(datos_ex[4])
            superadmin = bool(datos_ex[5])

    ok = bd.actualizar_usuario(username, nuevo_username, nuevo_rol, permisos, email, superadmin, bloqueado)
    if ok:
        flash(f' Usuario "{username}" actualizado correctamente.', 'success')
    else:
        flash(f' No se pudo actualizar el usuario "{username}" (si es el único superadmin, no puede ser des-promovido ni bloqueado).', 'error')

    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuario/<username>/desbloquear', methods=['POST'])
@login_requerido
def admin_usuario_desbloquear(username):
    """
    Desbloquea un usuario bloqueado por intentos fallidos.
    Solo superadmin puede usar esta ruta.
    """
    if not session.get('es_superadmin'):
        flash(' Solo el superadmin puede desbloquear usuarios.', 'error')
        return redirect(url_for('inicio'))

    ok = bd.desbloquear_usuario(username)
    if ok:
        flash(f' Usuario "{username}" desbloqueado correctamente.', 'success')
    else:
        flash(f' No se pudo desbloquear "{username}".', 'error')

    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuario/<username>/pass', methods=['POST'])
@login_requerido
def admin_usuario_pass(username):
    """
    Cambia la contraseña de un usuario.
    Requiere permiso usuarios:gestionar.
    """
    if not _puede("usuarios", "gestionar"):
        flash(' No tienes permiso para cambiar contraseñas.', 'error')
        return redirect(url_for('inicio'))

    # Evitar que no-superadmins cambien la contraseña de un superadmin
    if bd.es_superadmin(username) and not session.get('es_superadmin'):
        flash(' Solo el superadmin puede cambiar la contraseña de otro superadmin.', 'error')
        return redirect(url_for('admin_usuarios'))

    nueva_pass   = request.form.get('nueva_password', '').strip()
    confirmar    = request.form.get('confirmar_password', '').strip()

    if not nueva_pass or not confirmar:
        flash(' Ambos campos de contraseña son obligatorios.', 'error')
        return redirect(url_for('admin_usuarios'))

    if nueva_pass != confirmar:
        flash(' Las contraseñas no coinciden.', 'error')
        return redirect(url_for('admin_usuarios'))

    if not validar_complejidad_password(nueva_pass):
        flash(' La nueva contraseña no cumple con los requisitos de seguridad (mínimo 8 caracteres, mayúsculas, minúsculas, números y caracteres especiales).', 'error')
        return redirect(url_for('admin_usuarios'))

    ok = bd.actualizar_password_usuario(username, nueva_pass)
    if ok:
        flash(f' Contraseña de "{username}" actualizada.', 'success')
        bd.registrar_accion_auditoria(session.get('usuario'), 'Cambio Contraseña', f"Cambió contraseña del usuario '{username}'")
    else:
        flash(f' No se pudo cambiar la contraseña de "{username}".', 'error')

    return redirect(url_for('admin_usuarios'))


@app.route('/admin/usuario/<username>/borrar', methods=['POST'])
@login_requerido
def admin_usuario_borrar(username):
    """
    Elimina un usuario del sistema.
    Requiere permiso usuarios:gestionar. No permite eliminar al último superadmin.
    """
    if not _puede("usuarios", "gestionar"):
        flash(' No tienes permiso para eliminar usuarios.', 'error')
        return redirect(url_for('inicio'))

    # (Se eliminó la restricción de borrar admins por otros admins)

    ok = bd.eliminar_usuario(username)
    if ok:
        flash(f' Usuario "{username}" eliminado.', 'success')
    else:
        flash(f' No se pudo eliminar "{username}". '
              'Asegúrate de que no sea el único administrador.', 'error')

    return redirect(url_for('admin_usuarios'))


@app.route('/respaldo', methods=['GET', 'POST'])
@login_requerido
def generar_respaldo_manual():
    """Ruta para generar y descargar un respaldo manual de la BD (para todos los usuarios)."""
    filepath = crear_respaldo()
    if filepath and os.path.exists(filepath):
        flash('Respaldo generado correctamente.', 'success')
        bd.registrar_accion_auditoria(session.get('usuario'), 'Respaldo DB', 'Generó un respaldo manual de la base de datos.')
        return send_file(filepath, as_attachment=True)
    else:
        flash('Error al generar el respaldo de la base de datos.', 'error')
        return redirect(url_for('inicio'))

# =============================================================================
# DIAGNÓSTICO — Ver datos crudos desde el navegador
# =============================================================================

@app.route('/diagnostico')
@login_requerido
def diagnostico():
    """Muestra los datos crudos de los primeros 10 productos para depuración."""
    if not _puede("usuarios", "gestionar"):
        flash(' Solo administradores.', 'error')
        return redirect(url_for('inicio'))
    from src.database import ConexionBD as BD
    bd_local = BD()
    prods = bd_local.obtener_productos()
    html = '<h2>Diagnostico de datos (primeros 10 productos)</h2>'
    html += '<table border="1" cellpadding="6" style="border-collapse:collapse;font-family:monospace;font-size:13px;">'
    html += '<tr><th>#</th><th>sku [0]</th><th>nombre [1]</th><th>desc [2]</th><th>MARCA [3]</th><th>compat [4]</th><th>PRECIO [5]</th><th>cat [6]</th><th>exist [7]</th></tr>'
    for i, p in enumerate(prods[:10]):
        html += f'<tr><td>{i+1}</td>'
        for idx in range(min(8, len(p))):
            val = p[idx] if idx < len(p) else '—'
            clase = ' style="background:#fff3cd;"' if idx in (3, 5) else ''
            html += f'<td{clase}>{repr(val)}</td>'
        html += '</tr>'
    html += '</table>'
    html += f'<p>Total productos en BD: {len(prods)}</p>'
    html += '<p style="color:#666;">Las columnas MARCA [3] y PRECIO [5] estan resaltadas.</p>'
    return html


# =============================================================================
# MÓDULO GASTOS MARKETING
# =============================================================================

@app.route('/gastos')
@login_requerido
def ver_gastos():
    """
    Módulo Gastos Marketing — Visualización y registro de gastos.
    Filtra los gastos por mes y año. Si no se especifican, se usa el mes actual.
    """
    if not _puede("gastos", "ver"):
        flash(' No tienes permisos para ver el módulo de gastos.', 'error')
        return redirect(url_for('inicio'))

    import datetime
    hoy = datetime.date.today()
    
    # Obtener mes y año seleccionados (por defecto el actual)
    mes = request.args.get('mes', hoy.month, type=int)
    anio = request.args.get('anio', hoy.year, type=int)

    # Obtener gastos de la base de datos
    gastos_pub = bd.obtener_gastos_publicidad(mes, anio)
    gastos_lon = bd.obtener_gastos_lonas(mes, anio)

    # Calcular totales
    total_pub = sum(float(g[5]) for g in gastos_pub)
    total_lon = sum(float(g[5]) for g in gastos_lon)
    total_general = total_pub + total_lon

    # Meses disponibles para filtrar
    meses_bd = bd.obtener_meses_disponibles_gastos()
    
    # Asegurar que el mes/año actual o el seleccionado estén en la lista de meses para que el usuario pueda seleccionarlos
    meses_disponibles = []
    combo_actual = (mes, anio)
    combo_hoy = (hoy.month, hoy.year)
    
    # Convertir a set de tuplas para evitar duplicados
    set_meses = set((m[0], m[1]) for m in meses_bd)
    set_meses.add(combo_actual)
    set_meses.add(combo_hoy)
    
    # Nombre de los meses en español para la UI
    nombres_meses = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }

    # Ordenar los meses disponibles (año desc, mes desc)
    meses_disponibles = sorted(list(set_meses), key=lambda x: (x[1], x[0]), reverse=True)
    
    # Formatear la lista para el dropdown
    lista_meses_formateados = []
    for m, a in meses_disponibles:
        lista_meses_formateados.append({
            'mes': m,
            'anio': a,
            'nombre': f"{nombres_meses.get(m, 'Mes')} {a}"
        })

    return render_template(
        'gastos.html',
        gastos_publicidad=gastos_pub,
        gastos_lonas=gastos_lon,
        total_publicidad=total_pub,
        total_lonas=total_lon,
        total_general=total_general,
        mes_seleccionado=mes,
        anio_seleccionado=anio,
        meses_filtros=lista_meses_formateados,
        lista_aliados=bd.obtener_todos_aliados(),
        lista_clientes=bd.obtener_clientes()
    )


@app.route('/gastos/publicidad/nuevo', methods=['POST'])
@login_requerido
def nuevo_gasto_publicidad():
    """Registra un nuevo gasto de publicidad digital."""
    if not _puede("gastos", "gestionar"):
        flash(' No tienes permisos para gestionar gastos.', 'error')
        return redirect(url_for('ver_gastos'))

    post = request.form.get('post', '').strip()
    objetivo = request.form.get('objetivo', '').strip()
    metodo = request.form.get('metodo', '').strip()
    costo_dia = request.form.get('costo_dia', 0, type=float)
    total = request.form.get('total', 0, type=float)
    fecha_inicio = request.form.get('fecha_inicio', '').strip()
    fecha_fin = request.form.get('fecha_fin', '').strip()
    comentario = request.form.get('comentario', '').strip()
    aliado_id = None
    cliente_rif = request.form.get('cliente_rif', '').strip()

    if not post or not objetivo or not metodo or not fecha_inicio or not fecha_fin:
        flash(' Todos los campos (excepto comentario) son obligatorios para publicidad.', 'error')
        return redirect(url_for('ver_gastos'))

    # Si total es 0, calcularlo automáticamente a partir de los días de duración si es posible
    try:
        from datetime import datetime
        d_ini = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        d_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
        dias = (d_fin - d_ini).days + 1
        if dias > 0 and total == 0:
            total = costo_dia * dias
    except Exception:
        pass

    creado_por = session.get('usuario')
    ok = bd.crear_gasto_publicidad(
        post, objetivo, metodo, costo_dia, total, fecha_inicio, fecha_fin, 
        comentario, creado_por, aliado_id=aliado_id, cliente_rif=cliente_rif
    )

    if ok:
        flash(f' Gasto de publicidad "{post}" registrado correctamente.', 'exito')
        bd.registrar_accion_auditoria(creado_por, 'Registrar Gasto Publicidad', f'Gasto registrado: {post} ({metodo}) por ${total:.2f}')
    else:
        flash(' Error al guardar el gasto de publicidad en la base de datos.', 'error')

    try:
        from datetime import datetime
        dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        return redirect(url_for('ver_gastos', mes=dt.month, anio=dt.year))
    except Exception:
        return redirect(url_for('ver_gastos'))


@app.route('/gastos/publicidad/eliminar/<int:gasto_id>', methods=['POST'])
@login_requerido
def eliminar_gasto_publicidad(gasto_id):
    """Elimina un registro de gasto de publicidad."""
    if not _puede("gastos", "gestionar"):
        flash(' No tienes permisos para eliminar gastos.', 'error')
        return redirect(url_for('ver_gastos'))

    creado_por = session.get('usuario')
    ok = bd.eliminar_gasto_publicidad(gasto_id)

    if ok:
        flash(f' Gasto de publicidad #{gasto_id} eliminado.', 'exito')
        bd.registrar_accion_auditoria(creado_por, 'Eliminar Gasto Publicidad', f'Eliminó gasto de publicidad #{gasto_id}')
    else:
        flash(' Error al eliminar el gasto de publicidad.', 'error')

    return redirect(request.referrer or url_for('ver_gastos'))


@app.route('/gastos/publicidad/estadisticas/<int:gasto_id>', methods=['POST'])
@login_requerido
def guardar_estadisticas_publicidad(gasto_id):
    """Guarda o actualiza las estadísticas (alcance, clics, conversiones, ingresos) de una pauta."""
    if not _puede("gastos", "gestionar"):
        flash(' No tienes permisos para gestionar gastos.', 'error')
        return redirect(url_for('ver_gastos'))

    alcance_val = request.form.get('alcance', '').strip()
    clics_val = request.form.get('clics', '').strip()
    conversiones_val = request.form.get('conversiones', '').strip()
    ingresos_val = request.form.get('ingresos', '').strip()

    alcance = int(alcance_val) if alcance_val and alcance_val.isdigit() else None
    clics = int(clics_val) if clics_val and clics_val.isdigit() else None
    conversiones = int(conversiones_val) if conversiones_val and conversiones_val.isdigit() else None
    ingresos = float(ingresos_val) if ingresos_val else None

    creado_por = session.get('usuario')
    ok = bd.actualizar_estadisticas_publicidad(gasto_id, alcance, clics, conversiones, ingresos)

    if ok:
        flash(' Estadísticas de la pauta actualizadas correctamente.', 'exito')
        bd.registrar_accion_auditoria(creado_por, 'Actualizar Estadísticas Publicidad', f'Actualizó estadísticas del gasto publicitario #{gasto_id}')
    else:
        flash(' Error al actualizar las estadísticas en la base de datos.', 'error')

    return redirect(request.referrer or url_for('ver_gastos'))


@app.route('/gastos/lona/nuevo', methods=['POST'])
@login_requerido
def nuevo_gasto_lona():
    """Registra un nuevo gasto de lona física / insumos."""
    if not _puede("gastos", "gestionar"):
        flash(' No tienes permisos para gestionar gastos.', 'error')
        return redirect(url_for('ver_gastos'))

    herramienta = request.form.get('herramienta', '').strip()
    uso = request.form.get('uso', '').strip()
    precio = request.form.get('precio', 0, type=float)
    para_quien = request.form.get('para_quien', '').strip()
    total = request.form.get('total', 0, type=float)
    comentario = request.form.get('comentario', '').strip()
    
    # Nuevos campos de segmentación
    categoria = request.form.get('categoria', 'Otros').strip()
    aliado_rif = request.form.get('aliado_rif', '').strip()
    cantidad = request.form.get('cantidad', 1, type=int)

    # Validaciones y asignaciones según categoría
    if not herramienta or not uso:
        flash(' El concepto y el uso son campos obligatorios.', 'error')
        return redirect(url_for('ver_gastos'))

    aliado_id = None
    if categoria == 'Material para Aliado':
        if not aliado_rif:
            flash(' Debes seleccionar un cliente/aliado para esta categoría.', 'error')
            return redirect(url_for('ver_gastos'))
        # Recuperar o crear el aliado por su RIF
        cliente_datos = bd.obtener_cliente(aliado_rif)
        nombre_aliado = cliente_datos[1] if cliente_datos else aliado_rif
        
        # En database, necesitamos obtener el aliado_id o crearlo
        aliado_id_bd = bd.obtener_o_crear_aliado_por_rif(aliado_rif, nombre_aliado)
        if aliado_id_bd:
            aliado_id = aliado_id_bd
            para_quien = nombre_aliado
        else:
            para_quien = f"Aliado {aliado_rif}"
    else:
        # Para servicios, herramientas o POP se usa el campo de texto ingresado
        aliado_id = None
        if not para_quien:
            para_quien = 'Departamento'

    # Calcular total si no se ingresó
    if total == 0:
        total = precio * cantidad

    creado_por = session.get('usuario')
    ok = bd.crear_gasto_lona(
        herramienta=herramienta,
        uso=uso,
        precio=precio,
        para_quien=para_quien,
        total=total,
        comentario=comentario,
        creado_por=creado_por,
        categoria=categoria,
        aliado_id=aliado_id,
        cantidad=cantidad
    )

    if ok:
        flash(f' Gasto registrado correctamente en "{categoria}": {herramienta} por ${total:.2f}.', 'exito')
        bd.registrar_accion_auditoria(creado_por, 'Registrar Gasto Lona/Insumo', f'Gasto ({categoria}): {herramienta} para {para_quien} por ${total:.2f}')
    else:
        flash(' Error al guardar el gasto de insumos en la base de datos.', 'error')

    return redirect(url_for('ver_gastos'))


@app.route('/gastos/lona/eliminar/<int:gasto_id>', methods=['POST'])
@login_requerido
def eliminar_gasto_lona(gasto_id):
    """Elimina un registro de gasto de lona."""
    if not _puede("gastos", "gestionar"):
        flash(' No tienes permisos para eliminar gastos.', 'error')
        return redirect(url_for('ver_gastos'))

    creado_por = session.get('usuario')
    ok = bd.eliminar_gasto_lona(gasto_id)

    if ok:
        flash(f' Gasto de lona #{gasto_id} eliminado.', 'exito')
        bd.registrar_accion_auditoria(creado_por, 'Eliminar Gasto Lona', f'Eliminó gasto de lona #{gasto_id}')
    else:
        flash(' Error al eliminar el gasto de lona.', 'error')

    return redirect(request.referrer or url_for('ver_gastos'))


@app.route('/gastos/lona/<int:gasto_id>/pago/nuevo', methods=['POST'])
@login_requerido
def nuevo_pago_lona(gasto_id):
    """Registra un nuevo pago parcial o total a un gasto de lona/insumo."""
    if not _puede("gastos", "gestionar"):
        flash(' No tienes permisos para gestionar pagos.', 'error')
        return redirect(url_for('ver_gastos'))

    monto = request.form.get('monto', 0, type=float)
    metodo = request.form.get('metodo_pago', '').strip()
    referencia = request.form.get('referencia', '').strip()

    if monto <= 0 or not metodo:
        flash(' Monto y método de pago son obligatorios.', 'error')
        return redirect(url_for('ver_gastos'))

    creado_por = session.get('usuario')
    ok = bd.registrar_pago_lona(gasto_id, monto, metodo, referencia, creado_por)

    if ok:
        flash(f' Abono de ${monto:.2f} registrado correctamente.', 'exito')
        bd.registrar_accion_auditoria(creado_por, 'Registrar Abono Lona', f'Abono de ${monto:.2f} (Método: {metodo}) para gasto #{gasto_id}')
    else:
        flash(' Error al registrar el abono.', 'error')

    return redirect(request.referrer or url_for('ver_gastos'))


@app.route('/gastos/lona/<int:gasto_id>/pagos', methods=['GET'])
@login_requerido
def ver_pagos_lona(gasto_id):
    """Devuelve el historial de pagos de un gasto de lona (JSON)."""
    from flask import jsonify
    if not _puede("gastos", "ver"):
        return jsonify({'error': 'No tienes permisos para ver pagos.'}), 403
        
    pagos = bd.obtener_pagos_por_lona(gasto_id)
    lista = []
    for p in pagos:
        lista.append({
            'id': p[0],
            'monto': p[1],
            'metodo_pago': p[2],
            'referencia': p[3],
            'fecha_pago': p[4].strftime('%Y-%m-%d %H:%M') if p[4] else '',
            'registrado_por': p[5]
        })
    return jsonify(lista)


@app.route('/gastos/exportar')
@login_requerido
def exportar_gastos_excel():
    """Genera y descarga un reporte en Excel con los gastos del mes seleccionado."""
    if not _puede("gastos", "ver"):
        flash(' No tienes permisos para exportar gastos.', 'error')
        return redirect(url_for('inicio'))

    import datetime
    from datetime import date
    hoy = date.today()
    mes = request.args.get('mes', hoy.month, type=int)
    anio = request.args.get('anio', hoy.year, type=int)

    # Nombres de meses en español
    nombres_meses = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
        7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
    }
    nombre_mes = nombres_meses.get(mes, "Mes")

    # Obtener gastos de la base de datos
    gastos_pub = bd.obtener_gastos_publicidad(mes, anio)
    gastos_lon = bd.obtener_gastos_lonas(mes, anio)

    total_pub = sum(float(g[5]) for g in gastos_pub)
    total_lon = sum(float(g[5]) for g in gastos_lon)
    total_general = total_pub + total_lon

    # Creación de Libro Excel
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as OpenpyxlImage

    wb = openpyxl.Workbook()

    # Definición de estilos
    font_family = "Segoe UI"
    
    font_title = Font(name=font_family, size=16, bold=True, color="1E3A8A")
    font_subtitle = Font(name=font_family, size=11, italic=True, color="475569")
    font_section = Font(name=font_family, size=13, bold=True, color="1E293B")
    
    font_card_title = Font(name=font_family, size=9, bold=True, color="64748B")
    font_card_value = Font(name=font_family, size=16, bold=True, color="1E293B")
    
    font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    font_data = Font(name=font_family, size=10)
    font_total = Font(name=font_family, size=11, bold=True, color="000000")
    
    fill_pub_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_lon_header = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    fill_card_pub = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid") # light blue
    fill_card_lon = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid") # light green
    fill_card_tot = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid") # light purple
    
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    border_thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    border_card = Border(
        left=Side(style='medium', color='CBD5E1'),
        right=Side(style='medium', color='CBD5E1'),
        top=Side(style='medium', color='CBD5E1'),
        bottom=Side(style='medium', color='CBD5E1')
    )
    border_total = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    # ----------------- HOJA 1: RESUMEN -----------------
    ws1 = wb.active
    ws1.title = "Resumen de Gastos"
    ws1.views.sheetView[0].showGridLines = True

    # Insertar logotipo
    logo_inserted = False
    logo_path = os.path.join("almacen_activos", "Logo", "logo.png")
    if os.path.exists(logo_path):
        try:
            img = OpenpyxlImage(logo_path)
            # Redimensionar el logo para caber bien (alto 55px)
            aspect_ratio = img.width / img.height
            img.height = 55
            img.width = int(55 * aspect_ratio)
            ws1.add_image(img, "A2")
            logo_inserted = True
        except Exception as e:
            print(f" [Excel] Error al insertar logo: {e}")

    # Títulos y subtítulos
    col_idx = 3 if logo_inserted else 1
    
    ws1.cell(row=2, column=col_idx, value="IMPORTADORA UZIEL C.A.").font = font_title
    ws1.cell(row=3, column=col_idx, value=f"Reporte Mensual de Gastos de Marketing — {nombre_mes} {anio}").font = font_subtitle
    ws1.cell(row=4, column=col_idx, value=f"Generado el: {hoy.strftime('%d/%m/%Y')}").font = font_subtitle

    # Tarjetas de Resumen
    ws1.cell(row=7, column=1, value="RESUMEN EJECUTIVO DEL MES").font = font_section
    
    # Tarjeta 1: Publicidad Digital
    ws1.merge_cells("B9:C9")
    ws1.merge_cells("B10:C10")
    ws1.cell(row=9, column=2, value="PUBLICIDAD DIGITAL").font = font_card_title
    ws1.cell(row=9, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=10, column=2, value=total_pub).font = font_card_value
    ws1.cell(row=10, column=2).number_format = "$#,##0.00"
    ws1.cell(row=10, column=2).alignment = Alignment(horizontal="center", vertical="center")
    
    # Tarjeta 2: Lonas y Físicos
    ws1.merge_cells("E9:F9")
    ws1.merge_cells("E10:F10")
    ws1.cell(row=9, column=5, value="LONAS Y OTROS FÍSICOS").font = font_card_title
    ws1.cell(row=9, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=10, column=5, value=total_lon).font = font_card_value
    ws1.cell(row=10, column=5).number_format = "$#,##0.00"
    ws1.cell(row=10, column=5).alignment = Alignment(horizontal="center", vertical="center")

    # Tarjeta 3: Gran Total
    ws1.merge_cells("H9:I9")
    ws1.merge_cells("H10:I10")
    ws1.cell(row=9, column=8, value="GRAN TOTAL MENSUAL").font = Font(name=font_family, size=9, bold=True, color="6B21A8")
    ws1.cell(row=9, column=8).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=10, column=8, value=total_general).font = Font(name=font_family, size=16, bold=True, color="6B21A8")
    ws1.cell(row=10, column=8).number_format = "$#,##0.00"
    ws1.cell(row=10, column=8).alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar estilos a tarjetas (rellenos y bordes)
    for r in range(9, 11):
        for c in range(2, 4): # Col B, C
            cell = ws1.cell(row=r, column=c)
            cell.fill = fill_card_pub
            cell.border = border_card
        for c in range(5, 7): # Col E, F
            cell = ws1.cell(row=r, column=c)
            cell.fill = fill_card_lon
            cell.border = border_card
        for c in range(8, 10): # Col H, I
            cell = ws1.cell(row=r, column=c)
            cell.fill = fill_card_tot
            cell.border = border_card

    # Información
    ws1.cell(row=13, column=1, value="INFORMACIÓN DEL REPORTE").font = font_section
    ws1.cell(row=14, column=1, value="• Este reporte consolida la inversión publicitaria digital en redes/plataformas y los consumibles físicos.").font = font_data
    ws1.cell(row=15, column=1, value="• Puedes navegar en las pestañas inferiores para revisar el desglose y las métricas de rendimiento correspondientes.").font = font_data

    # Dimensiones de columnas Hoja 1
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 5
    ws1.column_dimensions['E'].width = 15
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 5
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 15

    # ----------------- HOJA 2: PUBLICIDAD DIGITAL -----------------
    ws2 = wb.create_sheet(title="Publicidad Digital")
    ws2.views.sheetView[0].showGridLines = True
    
    headers_pub = [
        "Post / Campaña", "Objetivo", "Plataforma / Canal", "Costo/Día ($)", "Total Campaña ($)", 
        "Fecha Inicio", "Fecha Fin", "Alcance", "Clics", "CTR", "Resultados", "CPA ($)", 
        "Ingresos ($)", "ROAS", "Cliente", "Comentario / Notas", "Creado Por"
    ]
    
    for c_num, header in enumerate(headers_pub, 1):
        cell = ws2.cell(row=1, column=c_num, value=header)
        cell.font = font_header
        cell.fill = fill_pub_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[1].height = 28

    # Filas de datos
    row_idx = 2
    for g in gastos_pub:
        # g = (id, post, objetivo, metodo, costo_dia, total, fecha_inicio, fecha_fin, comentario, creado_por, fecha_creacion, alcance, clics, conversiones, ingresos, aliado_id, nombre_aliado, cliente_rif, nombre_empresa)
        post_val = g[1]
        obj_val = g[2]
        met_val = g[3]
        costo_val = float(g[4])
        total_val = float(g[5])
        
        f_ini = g[6].strftime('%d/%m/%Y') if hasattr(g[6], 'strftime') else str(g[6])
        f_fin = g[7].strftime('%d/%m/%Y') if hasattr(g[7], 'strftime') else str(g[7])
        
        coment_val = g[8] if g[8] else ""
        por_val = g[9].capitalize() if g[9] else ""
        
        alcance_val = g[11]
        clics_val = g[12]
        
        ctr_val = None
        if alcance_val and clics_val and alcance_val > 0:
            ctr_val = clics_val / alcance_val
            
        conv_val = g[13]
        
        cpa_val = None
        if conv_val and conv_val > 0:
            cpa_val = total_val / conv_val
            
        ingr_val = g[14]
        
        roas_val = None
        if ingr_val is not None and total_val > 0:
            roas_val = float(ingr_val) / total_val
            
        # Relaciones
        nom_cliente = g[18] if len(g) > 18 and g[18] else ""

        row_data = [
            post_val, obj_val, met_val, costo_val, total_val, 
            f_ini, f_fin, alcance_val, clics_val, ctr_val, 
            conv_val, cpa_val, ingr_val, roas_val, nom_cliente,
            coment_val, por_val
        ]

        for col_num, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_num, value=val)
            cell.font = font_data
            cell.border = border_thin
            
            # Formatos y alineaciones
            if col_num in [1, 2, 3, 15, 16]: # Texto
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_num in [6, 7, 17]: # Fechas / Creador
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num in [4, 5, 12, 13]: # Moneda
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "$#,##0.00"
            elif col_num in [8, 9, 11]: # Enteros
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    cell.number_format = "#,##0"
            elif col_num == 10: # Porcentaje (CTR)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    cell.number_format = "0.0%"
            elif col_num == 14: # Factor ROAS
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if val is not None:
                    cell.number_format = '0.0"x"'
            
            # Colores alternos
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
        row_idx += 1

    # Fila de Totales
    total_row = row_idx
    ws2.cell(row=total_row, column=1, value="Total Publicidad").font = font_total
    ws2.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
    ws2.cell(row=total_row, column=1).border = border_total

    for c in range(1, len(headers_pub) + 1):
        ws2.cell(row=total_row, column=c).border = border_total
        ws2.cell(row=total_row, column=c).font = font_total

    ws2.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})").number_format = "$#,##0.00"
    ws2.cell(row=total_row, column=5).alignment = Alignment(horizontal="right")

    # Autoajuste de columnas
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                val_str = "$999,999.00"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ----------------- HOJA 3: LONAS Y FÍSICOS -----------------
    ws3 = wb.create_sheet(title="Lonas y Físicos")
    ws3.views.sheetView[0].showGridLines = True
    
    headers_lon = [
        "Categoría", "Concepto / Material", "Uso / Propósito", "Cantidad", 
        "Precio Unitario ($)", "Total ($)", "Destinatario", "Aliado Comercial", 
        "Comentario / Notas", "Creado Por"
    ]

    for c_num, header in enumerate(headers_lon, 1):
        cell = ws3.cell(row=1, column=c_num, value=header)
        cell.font = font_header
        cell.fill = fill_lon_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.row_dimensions[1].height = 28

    # Filas de datos
    row_idx = 2
    for g in gastos_lon:
        cat_val = g[9] if g[9] else "Otros"
        herr_val = g[1]
        uso_val = g[2]
        cant_val = int(g[11]) if g[11] is not None else 1
        prec_val = float(g[3])
        tot_val = float(g[5])
        dest_val = g[4]
        nom_aliado = g[12] if len(g) > 12 and g[12] else ""
        coment_val = g[6] if g[6] else ""
        por_val = g[7].capitalize() if g[7] else ""

        row_data = [
            cat_val, herr_val, uso_val, cant_val, 
            prec_val, tot_val, dest_val, nom_aliado, 
            coment_val, por_val
        ]

        for col_num, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_num, value=val)
            cell.font = font_data
            cell.border = border_thin
            
            if col_num in [1, 2, 3, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col_num == 10:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_num == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "#,##0"
            elif col_num in [5, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "$#,##0.00"
            
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
        row_idx += 1

    # Totales Lonas
    total_row = row_idx
    ws3.cell(row=total_row, column=1, value="Total Lonas y Físicos").font = font_total
    ws3.cell(row=total_row, column=1).alignment = Alignment(horizontal="left")
    ws3.cell(row=total_row, column=1).border = border_total

    for c in range(1, len(headers_lon) + 1):
        ws3.cell(row=total_row, column=c).border = border_total
        ws3.cell(row=total_row, column=c).font = font_total

    ws3.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row-1})").number_format = "$#,##0.00"
    ws3.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")

    for col in ws3.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                val_str = "$999,999.00"
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Reporte_Gastos_Marketing_{nombre_mes}_{anio}.xlsx"
    
    creado_por = session.get('usuario')
    bd.registrar_accion_auditoria(creado_por, 'Exportar Excel Gastos', f'Exportó gastos de {nombre_mes} {anio} a Excel ({filename})')

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# =============================================================================
# MÓDULO: IMPRESIÓN 3D
# =============================================================================

@app.route('/impresion3d')
@login_requerido
def impresion_3d():
    if not _puede("impresion3d", "ver"):
        flash(' No tienes permiso para ver el inventario 3D.', 'error')
        return redirect(url_for('inicio'))
    piezas = bd.obtener_piezas_3d()
    return render_template('impresion3d.html', piezas=piezas)

@app.route('/impresion3d/agregar', methods=['POST'])
@login_requerido
def agregar_pieza_3d():
    if not _puede("impresion3d", "agregar"):
        flash(' No tienes permiso para agregar piezas 3D.', 'error')
        return redirect(url_for('impresion_3d'))
    
    sku = request.form.get('sku', '').strip()
    nombre = request.form.get('nombre', '').strip()
    
    try:
        tiempo_min = int(request.form.get('tiempo_minutos', 0))
        peso_gramos = float(request.form.get('peso_gramos', 0.0))
        cantidad = int(request.form.get('cantidad', 1))
        
        costo_material = float(request.form.get('costo_material', 0.0))
        gasto_impresion = float(request.form.get('gasto_impresion', 0.0))
        costo_unitario = float(request.form.get('costo_unitario', 0.0))
        costo_total = float(request.form.get('costo_total', 0.0))
    except ValueError:
        flash('Error en los valores numéricos.', 'error')
        return redirect(url_for('impresion_3d'))
    
    if sku and nombre:
        exito = bd.agregar_pieza_3d(sku, nombre, tiempo_min, peso_gramos, costo_material, gasto_impresion, cantidad, costo_unitario, costo_total)
        if exito:
            flash(f' Pieza {nombre} guardada exitosamente.', 'exito')
        else:
            flash(' Error al guardar. Verifica que el SKU no esté duplicado.', 'error')
    else:
        flash(' Datos incompletos.', 'error')
        
    return redirect(url_for('impresion_3d'))

@app.route('/impresion3d/eliminar/<int:id_pieza>', methods=['POST'])
@login_requerido
def eliminar_pieza_3d(id_pieza):
    if not _puede("impresion3d", "eliminar"):
        flash(' No tienes permiso para eliminar piezas 3D.', 'error')
        return redirect(url_for('impresion_3d'))
    
    exito = bd.eliminar_pieza_3d(id_pieza)
    if exito:
        flash(' Pieza eliminada exitosamente.', 'exito')
    else:
        flash(' Error al eliminar la pieza.', 'error')
    return redirect(url_for('impresion_3d'))

# =============================================================================
# PUNTO DE ENTRADA (solo para desarrollo local)
# =============================================================================

if __name__ == '__main__':
    # NOTA: Para producción en Render, usar Gunicorn: gunicorn portal_web:app
    # El modo debug se activa solo si la variable FLASK_DEBUG=1 está definida,
    # para evitar su uso accidental en producción.
    modo_debug = os.getenv("FLASK_DEBUG", "0") == "1"
    app.run(host='0.0.0.0', port=5000, debug=modo_debug)
